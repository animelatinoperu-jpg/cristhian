import datetime as dt
from io import BytesIO
from unittest.mock import patch

from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from django.urls import reverse
from openpyxl import load_workbook

from productions.models import (
    Crew,
    Customer,
    NuqueraEntry,
    Product,
    ProductionOrder,
    TemplateVersion,
    User,
    Worker,
)
from productions.services.nuquera_tareo_report import (
    NuqueraTareoReportError,
    build_nuquera_tareo_xlsx,
)
from productions.services.nuquera_tareo_report_pdf import _prepare_tareo_xlsx_for_pdf


class NuqueraTareoReportTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user("manager", password="Secure-test-123")
        self.user.is_superuser = True
        self.user.is_staff = True
        self.user.first_name = "Daniela"
        self.user.last_name = "Rivas"
        self.user.save(update_fields=["is_superuser", "is_staff", "first_name", "last_name"])
        customer = Customer.objects.create(name="GLOBAL TOP FOOD PERÃš S.A.C.")
        product = Product.objects.create(code="RM-001", description="POTA A GRANEL")
        template = TemplateVersion.objects.create(
            code="PP-V1",
            file=SimpleUploadedFile("t.xlsm", b"x"),
            original_filename="t.xlsm",
            sha256="d" * 64,
            uploaded_by=self.user,
        )
        self.production = ProductionOrder.objects.create(
            number=300,
            plant_lot="1064PPF03082026",
            customer_lot="PPF03082026",
            customer=customer,
            process="POTA A GRANEL",
            main_product=product,
            reception_date=dt.date(2026, 8, 3),
            production_date=dt.date(2026, 8, 3),
            shift=ProductionOrder.Shift.DAY,
            template_version=template,
            created_by=self.user,
        )
        self.charles = Crew.objects.create(code="NUQ-01", name="CHARLES")
        self.beatriz = Worker.objects.create(
            internal_code="NUQ-W1",
            full_name="BETRIZ VARGAS",
            active=True,
        )
        self.angelica = Worker.objects.create(
            internal_code="NUQ-W2",
            full_name="ANGELICA SABINO",
            active=True,
        )
        NuqueraEntry.objects.create(
            production=self.production,
            responsible=self.user,
            observation="",
            date=self.production.reception_date,
            shift=ProductionOrder.Shift.DAY,
            crew=self.charles,
            worker=self.beatriz,
            process="NUCAS",
            weight_kg="25.82",
            start_time=dt.time(7, 0),
            end_time=dt.time(12, 0),
        )
        NuqueraEntry.objects.create(
            production=self.production,
            responsible=self.user,
            observation="",
            date=self.production.reception_date,
            shift=ProductionOrder.Shift.DAY,
            crew=self.charles,
            worker=self.angelica,
            process="NUCAS",
            weight_kg="25.88",
            start_time=dt.time(7, 0),
            end_time=dt.time(12, 0),
        )
        NuqueraEntry.objects.create(
            production=self.production,
            responsible=self.user,
            observation="",
            date=self.production.reception_date,
            shift=ProductionOrder.Shift.DAY,
            crew=self.charles,
            worker=self.beatriz,
            process="NUCAS",
            weight_kg="25.66",
            start_time=dt.time(7, 5),
            end_time=dt.time(12, 0),
        )

    def test_builds_workbook_with_nuqueras_sheets(self):
        payload = build_nuquera_tareo_xlsx(self.production)
        workbook = load_workbook(BytesIO(payload), data_only=False)

        self.assertEqual(
            workbook.sheetnames,
            [
                "NUQUERAS",
                " NUQUERAS omar ",
                "TAREO CHARLES",
                "TAREO PAZ",
                "TAREO OMAR",
                "TAREO CUADRILLA 4",
                "TAREO CUADRILLA 5",
            ],
        )
        main = workbook["NUQUERAS"]
        self.assertEqual(main["N1"].value, dt.datetime(2026, 8, 3))
        self.assertEqual(main["D5"].value, "Lunes")
        self.assertEqual(main["N5"].value, "NUCA")
        self.assertEqual(main["C6"].value, "CUADRILLA CHARLES")
        self.assertEqual(main["C7"].value, "BETRIZ VARGAS")
        self.assertEqual(main["D7"].value, "ANGELICA SABINO")
        self.assertEqual(main["C8"].value, 25.82)
        self.assertEqual(main["D8"].value, 25.88)
        self.assertEqual(main["C9"].value, 25.66)
        self.assertEqual(main["A8"].value, 1)
        self.assertEqual(main["A9"].value, 2)
        self.assertEqual(main["C38"].value, "=SUM(C8:C37)")
        self.assertEqual(main["B39"].value, "=SUM(B38:P38)")

        charles_tareo = workbook["TAREO CHARLES"]
        self.assertEqual(charles_tareo["D10"].value, "DANIELA RIVAS")
        self.assertEqual(charles_tareo["L10"].value, "CHARLES")
        self.assertEqual(charles_tareo["F8"].value, dt.datetime(2026, 8, 3))
        self.assertEqual(charles_tareo["C13"].value, "BETRIZ VARGAS")
        self.assertEqual(charles_tareo["C14"].value, "ANGELICA SABINO")
        self.assertEqual(charles_tareo["M13"].value, 51.48)

    def test_paz_block_is_kept_empty_when_unused(self):
        payload = build_nuquera_tareo_xlsx(self.production)
        workbook = load_workbook(BytesIO(payload), data_only=False)
        main = workbook["NUQUERAS"]
        self.assertEqual(main["C43"].value, "CUADRILLA PAZ")
        self.assertIsNone(main["C45"].value)
        omar = workbook[" NUQUERAS omar "]
        self.assertEqual(omar["C6"].value, "OMAR")
        self.assertIsNone(omar["C8"].value)

    def test_names_keep_original_case_like_template(self):
        worker = Worker.objects.create(
            internal_code="NUQ-W6",
            full_name="Beatriz Vargas",
            active=True,
        )
        NuqueraEntry.objects.create(
            production=self.production,
            responsible=self.user,
            date=self.production.reception_date,
            shift=ProductionOrder.Shift.DAY,
            crew=self.charles,
            worker=worker,
            process="NUCAS",
            weight_kg="10.00",
            start_time=dt.time(8, 0),
            end_time=dt.time(12, 0),
        )
        payload = build_nuquera_tareo_xlsx(self.production)
        workbook = load_workbook(BytesIO(payload), data_only=False)
        main = workbook["NUQUERAS"]
        self.assertEqual(main["C6"].value, "CUADRILLA CHARLES")
        self.assertEqual(main["C7"].value, "BETRIZ VARGAS")
        self.assertEqual(main["D7"].value, "ANGELICA SABINO")
        self.assertEqual(main["E7"].value, "Beatriz Vargas")
        charles_tareo = workbook["TAREO CHARLES"]
        self.assertEqual(charles_tareo["C13"].value, "BETRIZ VARGAS")
        self.assertEqual(charles_tareo["C14"].value, "ANGELICA SABINO")
        self.assertEqual(charles_tareo["C15"].value, "Beatriz Vargas")

    def test_more_than_three_crews_uses_dynamic_extra_block(self):
        omar_crew = Crew.objects.create(code="NUQ-02", name="OMAR")
        paz_crew = Crew.objects.create(code="NUQ-03", name="PAZ")
        omar_worker = Worker.objects.create(
            internal_code="NUQ-W3",
            full_name="MARTINA QUISPE",
            active=True,
        )
        paz_worker = Worker.objects.create(
            internal_code="NUQ-W4",
            full_name="SARA CORI",
            active=True,
        )
        for crew, worker in ((omar_crew, omar_worker), (paz_crew, paz_worker)):
            NuqueraEntry.objects.create(
                production=self.production,
                responsible=self.user,
                date=self.production.reception_date,
                shift=ProductionOrder.Shift.DAY,
                crew=crew,
                worker=worker,
                process="NUCAS",
                weight_kg="10.00",
                start_time=dt.time(8, 0),
                end_time=dt.time(12, 0),
            )
        another_crew = Crew.objects.create(code="NUQ-99", name="SALVAJES")
        another_worker = Worker.objects.create(
            internal_code="NUQ-W5",
            full_name="OTRO TRABAJADOR",
        )
        NuqueraEntry.objects.create(
            production=self.production,
            responsible=self.user,
            date=self.production.reception_date,
            shift=ProductionOrder.Shift.DAY,
            crew=another_crew,
            worker=another_worker,
            process="NUCAS",
            weight_kg="10.00",
            start_time=dt.time(8, 0),
            end_time=dt.time(12, 0),
        )
        payload = build_nuquera_tareo_xlsx(self.production)
        workbook = load_workbook(BytesIO(payload), data_only=False)
        self.assertIn("TAREO CUADRILLA 4", workbook.sheetnames)
        main = workbook["NUQUERAS"]
        self.assertEqual(main["C76"].value, "CUADRILLA SALVAJES")
        self.assertEqual(main["C77"].value, "OTRO TRABAJADOR")
        self.assertEqual(main["C78"].value, 10.00)
        self.assertEqual(main["A78"].value, 1)
        self.assertEqual(main["C108"].value, "=SUM(C78:C107)")
        self.assertEqual(main["B109"].value, "=SUM(B108:P108)")
        tareo = workbook["TAREO CUADRILLA 4"]
        self.assertEqual(tareo["L10"].value, "SALVAJES")
        self.assertEqual(tareo["C13"].value, "OTRO TRABAJADOR")

    def test_six_crews_uses_pre_edited_and_dynamic_extra_blocks(self):
        entries = []
        for crew_index, crew_name in enumerate(("PAZ", "OMAR", "ALFA", "BETA", "GAMMA"), start=2):
            crew = Crew.objects.create(code=f"NUQ-{crew_index:02d}", name=crew_name)
            worker = Worker.objects.create(
                internal_code=f"NUQ-W10{crew_index}",
                full_name=f"TRABAJADOR {crew_name}",
                active=True,
            )
            entries.append((crew, worker))
        for crew, worker in entries:
            NuqueraEntry.objects.create(
                production=self.production,
                responsible=self.user,
                date=self.production.reception_date,
                shift=ProductionOrder.Shift.DAY,
                crew=crew,
                worker=worker,
                process="NUCAS",
                weight_kg="12.50",
                start_time=dt.time(8, 0),
                end_time=dt.time(12, 0),
            )
        payload = build_nuquera_tareo_xlsx(self.production)
        workbook = load_workbook(BytesIO(payload), data_only=False)
        for sheet in ("TAREO CUADRILLA 4", "TAREO CUADRILLA 5", "TAREO CUADRILLA 6"):
            self.assertIn(sheet, workbook.sheetnames)
        main = workbook["NUQUERAS"]
        expected = {
            76: ("ALFA", "TRABAJADOR ALFA", 77, 108, 109, "TAREO CUADRILLA 4"),
            113: ("BETA", "TRABAJADOR BETA", 114, 145, 146, "TAREO CUADRILLA 5"),
            150: ("GAMMA", "TRABAJADOR GAMMA", 151, 182, 183, "TAREO CUADRILLA 6"),
        }
        for dest, (label, worker_name, names_row, subtotal_row, total_row, tareo_name) in expected.items():
            self.assertEqual(main[f"C{dest}"].value, f"CUADRILLA {label}")
            self.assertEqual(main[f"C{names_row}"].value, worker_name)
            self.assertEqual(main[f"C{names_row + 1}"].value, 12.50)
            self.assertEqual(main[f"A{names_row + 1}"].value, 1)
            self.assertEqual(
                main[f"C{subtotal_row}"].value,
                f"=SUM(C{names_row + 1}:C{dest + 31})",
            )
            self.assertEqual(
                main[f"B{total_row}"].value,
                f"=SUM(B{subtotal_row}:P{subtotal_row})",
            )
            tareo = workbook[tareo_name]
            self.assertEqual(tareo["L10"].value, label)

    def test_download_view_requires_login_and_returns_xlsx(self):
        reverse_name = reverse("productions:nuquera_tareo_xlsx", args=[self.production.pk])
        response = self.client.get(reverse_name)
        self.assertNotEqual(response.status_code, 200)

        self.client.force_login(self.user)
        response = self.client.get(reverse_name)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("NUQUERAS_TAREO", response["Content-Disposition"])

        workbook = load_workbook(BytesIO(response.content), data_only=False)
        main = workbook["NUQUERAS"]
        self.assertEqual(main["C8"].value, 25.82)

    def test_prepare_tareo_xlsx_for_pdf_sets_print_areas(self):
        payload = build_nuquera_tareo_xlsx(self.production)
        pdf_ready = load_workbook(
            BytesIO(_prepare_tareo_xlsx_for_pdf(payload)),
            data_only=False,
        )
        self.assertEqual(
            str(pdf_ready["PESOS CHARLES"].print_area),
            "'PESOS CHARLES'!$A$1:$P$39",
        )
        self.assertEqual(
            str(pdf_ready["TAREO CHARLES"].print_area),
            "'TAREO CHARLES'!$A$1:$P$50",
        )
        for sheet_name in ("PESOS CHARLES", "TAREO CHARLES"):
            ws = pdf_ready[sheet_name]
            self.assertTrue(ws.sheet_properties.pageSetUpPr.fitToPage)
            self.assertEqual(ws.page_setup.fitToWidth, 1)
            self.assertEqual(ws.page_setup.orientation, "portrait")
            for row in ws.iter_rows(min_row=1, max_row=5):
                for cell in row:
                    if cell.font:
                        self.assertNotEqual(cell.font.name, "Arial Narrow")
        self.assertEqual(pdf_ready["PESOS CHARLES"].page_setup.fitToHeight, 0)
        self.assertEqual(pdf_ready["TAREO CHARLES"].page_setup.fitToHeight, 1)

    def test_prepare_tareo_xlsx_for_pdf_hides_empty_sheets(self):
        payload = build_nuquera_tareo_xlsx(self.production)
        pdf_ready = load_workbook(
            BytesIO(_prepare_tareo_xlsx_for_pdf(payload)),
            data_only=False,
        )
        for sheet_name in ("PESOS CHARLES", "TAREO CHARLES"):
            self.assertEqual(pdf_ready[sheet_name].sheet_state, "visible")
            self.assertIsNotNone(pdf_ready[sheet_name].print_area)
        for sheet_name in ("NUQUERAS", " NUQUERAS omar ", "TAREO PAZ", "TAREO OMAR"):
            self.assertEqual(pdf_ready[sheet_name].sheet_state, "hidden")
            self.assertIn(str(pdf_ready[sheet_name].print_area), ("", "None"))

    def test_pdf_creates_one_weight_sheet_per_active_crew_in_order(self):
        paz = Crew.objects.create(code="NUQ-04", name="PAZ")
        sara = Worker.objects.create(
            internal_code="NUQ-W7",
            full_name="SARA CORI",
            active=True,
        )
        NuqueraEntry.objects.create(
            production=self.production,
            responsible=self.user,
            date=self.production.reception_date,
            shift=ProductionOrder.Shift.DAY,
            crew=paz,
            worker=sara,
            process="NUCAS",
            weight_kg="30.00",
            start_time=dt.time(8, 0),
            end_time=dt.time(12, 0),
        )
        payload = build_nuquera_tareo_xlsx(self.production)
        pdf_ready = load_workbook(
            BytesIO(_prepare_tareo_xlsx_for_pdf(payload)),
            data_only=False,
        )
        self.assertEqual(
            pdf_ready.sheetnames[:4],
            ["PESOS CHARLES", "TAREO CHARLES", "PESOS PAZ", "TAREO PAZ"],
        )
        self.assertNotIn("PESOS OMAR", pdf_ready.sheetnames)

        pesos_charles = pdf_ready["PESOS CHARLES"]
        self.assertEqual(pesos_charles["C8"].value, 25.82)
        self.assertEqual(pesos_charles["C6"].value, "CUADRILLA CHARLES")
        self.assertIsNone(pesos_charles["C45"].value)

        pesos_paz = pdf_ready["PESOS PAZ"]
        self.assertEqual(pesos_paz["C45"].value, 30.00)
        self.assertEqual(pesos_paz["B45"].value, 1)
        self.assertIsNone(pesos_paz["C8"].value)
        self.assertTrue(pesos_paz.row_dimensions[6].hidden)
        self.assertEqual(
            str(pesos_paz.print_area),
            "'PESOS PAZ'!$A$1:$P$73",
        )

    def test_pdf_single_crew_only_shows_that_crew(self):
        payload = build_nuquera_tareo_xlsx(self.production)
        pdf_ready = load_workbook(
            BytesIO(_prepare_tareo_xlsx_for_pdf(payload)),
            data_only=False,
        )
        self.assertEqual(
            pdf_ready.sheetnames[:2],
            ["PESOS CHARLES", "TAREO CHARLES"],
        )
        self.assertNotIn("PESOS PAZ", pdf_ready.sheetnames)
        self.assertNotIn("PESOS OMAR", pdf_ready.sheetnames)
        self.assertEqual(pdf_ready["PESOS CHARLES"].sheet_state, "visible")
        self.assertEqual(pdf_ready["TAREO PAZ"].sheet_state, "hidden")

    def test_prepare_tareo_xlsx_for_pdf_fits_crew_name_row(self):
        payload = build_nuquera_tareo_xlsx(self.production)
        pdf_ready = load_workbook(
            BytesIO(_prepare_tareo_xlsx_for_pdf(payload)),
            data_only=False,
        )
        ws = pdf_ready["PESOS CHARLES"]
        self.assertEqual(ws["C7"].value, "BETRIZ VARGAS")
        self.assertEqual(ws["D7"].value, "ANGELICA SABINO")
        for column in ("C", "D"):
            alignment = ws[f"{column}7"].alignment
            self.assertTrue(alignment.wrap_text)
            self.assertEqual(alignment.vertical, "center")
            self.assertEqual(alignment.horizontal, "center")
        row = ws.row_dimensions[7]
        self.assertIsNone(row.height)
        self.assertFalse(row.customHeight)

    def test_pdf_dynamic_crew_gets_weight_sheet_from_extra_block(self):
        for crew_index, crew_name in ((2, "PAZ"), (3, "OMAR"), (4, "SALVAJES")):
            crew = Crew.objects.create(code=f"NUQ-{crew_index:02d}", name=crew_name)
            worker = Worker.objects.create(
                internal_code=f"NUQ-W30{crew_index}",
                full_name=f"TRABAJADOR {crew_name}",
                active=True,
            )
            NuqueraEntry.objects.create(
                production=self.production,
                responsible=self.user,
                date=self.production.reception_date,
                shift=ProductionOrder.Shift.DAY,
                crew=crew,
                worker=worker,
                process="NUCAS",
                weight_kg="20.00",
                start_time=dt.time(8, 0),
                end_time=dt.time(12, 0),
            )
        payload = build_nuquera_tareo_xlsx(self.production)
        pdf_ready = load_workbook(
            BytesIO(_prepare_tareo_xlsx_for_pdf(payload)),
            data_only=False,
        )
        self.assertEqual(
            pdf_ready.sheetnames[:8],
            [
                "PESOS CHARLES",
                "TAREO CHARLES",
                "PESOS PAZ",
                "TAREO PAZ",
                "PESOS OMAR",
                "TAREO OMAR",
                "PESOS SALVAJES",
                "TAREO CUADRILLA 4",
            ],
        )
        pesos_salvajes = pdf_ready["PESOS SALVAJES"]
        self.assertEqual(pesos_salvajes["C76"].value, "CUADRILLA SALVAJES")
        self.assertEqual(pesos_salvajes["C77"].value, "TRABAJADOR SALVAJES")
        self.assertEqual(pesos_salvajes["C78"].value, 20.00)
        self.assertEqual(
            str(pesos_salvajes.print_area),
            "'PESOS SALVAJES'!$A$1:$P$109",
        )
        alignment = pesos_salvajes["C77"].alignment
        self.assertTrue(alignment.wrap_text)
        self.assertEqual(alignment.vertical, "center")
        self.assertEqual(alignment.horizontal, "center")

    def test_pdf_six_crews_prints_all_extra_tareos(self):
        for crew_index, crew_name in enumerate(("PAZ", "OMAR", "ALFA", "BETA", "GAMMA"), start=2):
            crew = Crew.objects.create(code=f"NUQ-{crew_index:02d}", name=crew_name)
            worker = Worker.objects.create(
                internal_code=f"NUQ-W20{crew_index}",
                full_name=f"TRABAJADOR {crew_name}",
                active=True,
            )
            NuqueraEntry.objects.create(
                production=self.production,
                responsible=self.user,
                date=self.production.reception_date,
                shift=ProductionOrder.Shift.DAY,
                crew=crew,
                worker=worker,
                process="NUCAS",
                weight_kg="5.00",
                start_time=dt.time(8, 0),
                end_time=dt.time(12, 0),
            )
        payload = build_nuquera_tareo_xlsx(self.production)
        pdf_ready = load_workbook(
            BytesIO(_prepare_tareo_xlsx_for_pdf(payload)),
            data_only=False,
        )
        self.assertEqual(
            pdf_ready.sheetnames[:10],
            [
                "PESOS CHARLES",
                "TAREO CHARLES",
                "PESOS PAZ",
                "TAREO PAZ",
                "PESOS OMAR",
                "TAREO OMAR",
                "PESOS ALFA",
                "TAREO CUADRILLA 4",
                "PESOS BETA",
                "TAREO CUADRILLA 5",
            ],
        )
        self.assertEqual(pdf_ready["PESOS ALFA"]["C76"].value, "CUADRILLA ALFA")
        self.assertEqual(pdf_ready["PESOS BETA"]["C113"].value, "CUADRILLA BETA")
        self.assertEqual(
            str(pdf_ready["PESOS BETA"].print_area),
            "'PESOS BETA'!$A$1:$P$146",
        )

    def test_download_page_shows_pdf_button(self):
        self.client.force_login(self.user)
        page = self.client.get(reverse("productions:nuquera_create", args=[self.production.pk]))
        self.assertContains(page, "Descargar tareo de nucas Excel")
        self.assertContains(page, "Descargar tareo de nucas PDF")
        self.assertContains(page, "Agregar un trabajador y su cuadrilla")

    def test_quick_create_worker_and_crew(self):
        self.client.force_login(self.user)
        url = reverse("productions:nuquera_worker_quick_create", args=[self.production.pk])
        response = self.client.post(url, {"name": "   rosa   flores  ", "crew": "las estrellas"})
        self.assertEqual(response.status_code, 302)
        worker = Worker.objects.get(internal_code__startswith="NUQ-W", full_name="ROSA FLORES")
        self.assertTrue(worker.active)
        self.assertEqual(worker.position, "Nuquera")
        crew = Crew.objects.get(code__startswith="NUQ-", name="LAS ESTRELLAS")
        self.assertTrue(crew.active)
        self.assertEqual(worker.crew_id, crew.pk)
        self.assertTrue(worker.internal_code.startswith("NUQ-W"))
        self.assertTrue(crew.code.startswith("NUQ-"))

    def test_quick_create_reuses_existing_worker_and_crew(self):
        self.client.force_login(self.user)
        url = reverse("productions:nuquera_worker_quick_create", args=[self.production.pk])
        worker_count = Worker.objects.count()
        crew_count = Crew.objects.count()
        response = self.client.post(url, {"name": "  betriz vargas ", "crew": "charles"})
        self.assertEqual(response.status_code, 302)
        self.assertEqual(Worker.objects.count(), worker_count)
        self.assertEqual(Crew.objects.count(), crew_count)
        self.assertEqual(Worker.objects.filter(full_name="BETRIZ VARGAS").count(), 1)

    def test_quick_created_worker_appears_in_nuquera_form(self):
        self.client.force_login(self.user)
        url = reverse("productions:nuquera_worker_quick_create", args=[self.production.pk])
        self.client.post(url, {"name": "ROSA FLORES", "crew": "LAS ESTRELLAS"})
        page = self.client.get(reverse("productions:nuquera_create", args=[self.production.pk]))
        self.assertContains(page, "ROSA FLORES")
        self.assertContains(page, "LAS ESTRELLAS")

    def test_quick_create_requires_login(self):
        url = reverse("productions:nuquera_worker_quick_create", args=[self.production.pk])
        response = self.client.post(url, {"name": "ROSA FLORES", "crew": "LAS ESTRELLAS"})
        self.assertNotEqual(response.status_code, 200)

    def test_quick_create_blocked_when_approved(self):
        self.production.status = ProductionOrder.Status.APPROVED
        self.production.save(update_fields=["status"])
        self.client.force_login(self.user)
        url = reverse("productions:nuquera_worker_quick_create", args=[self.production.pk])
        response = self.client.post(url, {"name": "ROSA FLORES", "crew": "LAS ESTRELLAS"})
        self.assertEqual(response.status_code, 403)

    @patch("productions.views.build_nuquera_tareo_pdf", return_value=b"%PDF-1.4\n%%EOF")
    def test_pdf_download_uses_the_nuquera_tareo_endpoint(self, mocked_builder):
        self.client.force_login(self.user)
        response = self.client.get(
            reverse("productions:nuquera_tareo_pdf", args=[self.production.pk])
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/pdf")
        self.assertTrue(response.content.startswith(b"%PDF"))
        self.assertIn("NUQUERAS_TAREO_PP_300", response["Content-Disposition"])
        mocked_builder.assert_called_once_with(self.production)

    def test_crew_mode_shows_worker_buttons_and_locks_crew(self):
        self.client.force_login(self.user)
        url = (
            reverse("productions:nuquera_create", args=[self.production.pk])
            + f"?crew={self.charles.pk}"
        )
        page = self.client.get(url)
        self.assertEqual(page.status_code, 200)
        self.assertContains(page, "Elegir otra cuadrilla")
        self.assertContains(page, "BETRIZ VARGAS")
        self.assertContains(page, "ANGELICA SABINO")
        form = page.context["form"]
        self.assertTrue(form.fields["crew"].disabled)
        self.assertEqual(
            {w.pk for w in form.fields["worker"].queryset},
            {self.beatriz.pk, self.angelica.pk},
        )

    def test_worker_param_preselects_and_copies_last_entry(self):
        self.client.force_login(self.user)
        url = (
            reverse("productions:nuquera_create", args=[self.production.pk])
            + f"?crew={self.charles.pk}&worker={self.beatriz.pk}"
        )
        page = self.client.get(url)
        self.assertEqual(page.status_code, 200)
        initial = page.context["form"].initial
        self.assertEqual(initial.get("crew"), self.charles.pk)
        self.assertEqual(initial.get("worker"), self.beatriz.pk)
        self.assertEqual(initial.get("start_time"), dt.time(7, 5))
        self.assertEqual(initial.get("end_time"), dt.time(12, 0))
        self.assertEqual(initial.get("process"), "NUCAS")
        self.assertEqual(initial.get("shift"), ProductionOrder.Shift.DAY)

    def test_worker_param_defaults_hours_when_no_last_entry(self):
        self.client.force_login(self.user)
        newbie = Worker.objects.create(
            internal_code="NUQ-W9",
            full_name="NUEVA PERSONA",
            active=True,
            crew=self.charles,
        )
        url = (
            reverse("productions:nuquera_create", args=[self.production.pk])
            + f"?crew={self.charles.pk}&worker={newbie.pk}"
        )
        page = self.client.get(url)
        self.assertEqual(page.status_code, 200)
        initial = page.context["form"].initial
        self.assertEqual(initial.get("start_time"), dt.time(6, 0))
        self.assertEqual(initial.get("end_time"), dt.time(18, 0))
        self.assertEqual(initial.get("process"), self.production.process)
        self.assertEqual(initial.get("shift"), self.production.shift)

    def test_worker_param_focuses_form_on_weight_only(self):
        self.client.force_login(self.user)
        url = (
            reverse("productions:nuquera_create", args=[self.production.pk])
            + f"?crew={self.charles.pk}&worker={self.beatriz.pk}"
        )
        page = self.client.get(url)
        self.assertEqual(page.status_code, 200)
        self.assertTrue(page.context["nuquera_focused_worker"])
        content = page.content.decode("utf-8")
        self.assertIn("Registrando peso de BETRIZ VARGAS", content)
        self.assertIn("Guardar peso", content)
        for field_id in ("id_shift", "id_crew", "id_worker", "id_process", "id_start_time", "id_end_time", "id_observation"):
            self.assertIn(field_id, content)
        self.assertIn('id="id_weight_kg"', content)
        self.assertIn("d-none", content)

    def test_form_hidden_before_choosing_crew(self):
        self.client.force_login(self.user)
        page = self.client.get(reverse("productions:nuquera_create", args=[self.production.pk]))
        self.assertEqual(page.status_code, 200)
        self.assertContains(page, "Ver trabajadores")
        self.assertContains(page, "CUADRILLA DE TRABAJO")

    def test_form_visible_when_crew_selected(self):
        self.client.force_login(self.user)
        url = (
            reverse("productions:nuquera_create", args=[self.production.pk])
            + f"?crew={self.charles.pk}"
        )
        page = self.client.get(url)
        self.assertEqual(page.status_code, 200)
        self.assertContains(page, "Elegir otra cuadrilla")

    def test_crew_suggestions_only_include_this_production(self):
        self.client.force_login(self.user)
        other_crew = Crew.objects.create(code="NUQ-99", name="CUADRILLA DE AYER")
        Worker.objects.create(
            internal_code="NUQ-W8",
            full_name="ALGUIEN DE AYER",
            active=True,
            crew=other_crew,
        )
        page = self.client.get(reverse("productions:nuquera_create", args=[self.production.pk]))
        suggestions = page.context["nuquera_crew_suggestions"]
        self.assertIn((self.charles.pk, self.charles.name), suggestions)
        self.assertNotIn((other_crew.pk, other_crew.name), suggestions)

    def test_quick_created_crew_appears_immediately(self):
        self.client.force_login(self.user)
        url = reverse("productions:nuquera_worker_quick_create", args=[self.production.pk])
        self.client.post(url, {"name": "ROSA FLORES", "crew": "LAS ESTRELLAS"})
        page = self.client.get(reverse("productions:nuquera_create", args=[self.production.pk]))
        suggestions = page.context["nuquera_crew_suggestions"]
        self.assertTrue(any(name == "LAS ESTRELLAS" for _, name in suggestions))

    def test_crew_panel_shows_message_when_no_crews(self):
        self.client.force_login(self.user)
        fresh = ProductionOrder.objects.create(
            number=301,
            plant_lot="1064PPF03112026",
            customer_lot="PPF03112026",
            customer=self.production.customer,
            process="POTA A GRANEL",
            main_product=self.production.main_product,
            reception_date=dt.date(2026, 8, 4),
            production_date=dt.date(2026, 8, 4),
            shift=ProductionOrder.Shift.DAY,
            template_version=self.production.template_version,
            created_by=self.user,
        )
        page = self.client.get(reverse("productions:nuquera_create", args=[fresh.pk]))
        self.assertEqual(page.status_code, 200)
        self.assertContains(page, "Este parte todavía no tiene cuadrillas")
        self.assertEqual(page.context["nuquera_crew_suggestions"], [])

    def test_invalid_crew_and_worker_params_are_ignored(self):
        self.client.force_login(self.user)
        url = (
            reverse("productions:nuquera_create", args=[self.production.pk])
            + "?crew=99999&worker=99999"
        )
        page = self.client.get(url)
        self.assertEqual(page.status_code, 200)
        form = page.context["form"]
        self.assertFalse(form.fields["crew"].disabled)
        self.assertIsNone(form.initial.get("worker"))
        self.assertNotContains(page, "Elegir otra cuadrilla")

    def test_save_redirects_back_to_crew_mode(self):
        self.client.force_login(self.user)
        url = reverse("productions:nuquera_create", args=[self.production.pk])
        response = self.client.post(
            url,
            {
                "shift": ProductionOrder.Shift.DAY,
                "crew": str(self.charles.pk),
                "worker": str(self.beatriz.pk),
                "process": "NUCAS",
                "weight_kg": "30.00",
                "start_time": "08:00",
                "end_time": "12:00",
                "observation": "",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertEqual(
            response["Location"],
            f"{url}?crew={self.charles.pk}#operational-entry-form",
        )

    def test_crew_panel_hidden_when_approved(self):
        self.production.status = ProductionOrder.Status.APPROVED
        self.production.save(update_fields=["status"])
        self.client.force_login(self.user)
        url = (
            reverse("productions:nuquera_create", args=[self.production.pk])
            + f"?crew={self.charles.pk}"
        )
        page = self.client.get(url)
        self.assertEqual(page.status_code, 200)
        self.assertNotContains(page, "Ver trabajadores")
        self.assertNotContains(page, "Elegir otra cuadrilla")

    def test_crew_panel_hidden_when_editing(self):
        self.client.force_login(self.user)
        entry = NuqueraEntry.objects.filter(production=self.production).first()
        url = reverse(
            "productions:operational_entry_update",
            args=[self.production.pk, "nuqueras", entry.pk],
        )
        page = self.client.get(url)
        self.assertEqual(page.status_code, 200)
        self.assertNotContains(page, "Ver trabajadores")
        self.assertNotContains(page, "Elegir otra cuadrilla")

