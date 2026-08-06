import datetime as dt
import json
from io import BytesIO
from unittest.mock import patch

from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from django.urls import reverse
from openpyxl import load_workbook

from productions.models import (
    Crew,
    Customer,
    Product,
    ProductionOrder,
    TemplateVersion,
    TroqueladoEntry,
    User,
    Worker,
)
from productions.services.troquelado_report import (
    TroqueladoReportError,
    build_troquelado_xlsx,
)
from productions.services.troquelado_report_pdf import _prepare_troquelado_xlsx_for_pdf


class TroqueladoReportTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user("manager", password="Secure-test-123")
        self.user.is_superuser = True
        self.user.is_staff = True
        self.user.first_name = "Daniela"
        self.user.last_name = "Rivas"
        self.user.save(update_fields=["is_superuser", "is_staff", "first_name", "last_name"])
        customer = Customer.objects.create(name="GLOBAL TOP FOOD PERÃš S.A.C.")
        product = Product.objects.create(code="RM-001", description="POTA A GRANEL")
        template = TemplateVersion.objects.create(
            code="PP-V1",
            file=SimpleUploadedFile("t.xlsm", b"x"),
            original_filename="t.xlsm",
            sha256="d" * 64,
            uploaded_by=self.user,
        )
        self.production = ProductionOrder.objects.create(
            number=300,
            plant_lot="1064PPF03082026",
            customer_lot="PPF03082026",
            customer=customer,
            process="POTA A GRANEL",
            main_product=product,
            reception_date=dt.date(2026, 8, 3),
            production_date=dt.date(2026, 8, 3),
            shift=ProductionOrder.Shift.DAY,
            template_version=template,
            created_by=self.user,
        )
        self.crew = Crew.objects.create(code="TROQ-01", name="CHARLES")
        self.juan = Worker.objects.create(
            internal_code="TROQ-W1",
            full_name="JUAN PEREZ",
            active=True,
            crew=self.crew,
        )
        self.luis = Worker.objects.create(
            internal_code="TROQ-W2",
            full_name="LUIS ROJAS",
            active=True,
            crew=self.crew,
        )
        self.pedro = Worker.objects.create(
            internal_code="TROQ-W3",
            full_name="PEDRO TORRES",
            active=True,
            crew=self.crew,
        )

    def _entry(self, worker, product_type, cajas, kg, start, end, crew=None):
        return TroqueladoEntry.objects.create(
            production=self.production,
            responsible=self.user,
            date=self.production.reception_date,
            shift=ProductionOrder.Shift.DAY,
            crew=crew or self.crew,
            worker=worker,
            product_type=product_type,
            cajas=cajas,
            kg_por_caja=kg,
            weight_kg=round(cajas * kg, 2),
            start_time=start,
            end_time=end,
        )

    def _load_sample(self):
        self._entry(self.juan, "ANILLAS BLANCAS", 4, 20, dt.time(7, 0), dt.time(12, 0))
        self._entry(self.luis, "ANILLAS BLANCAS", 3, 20, dt.time(7, 0), dt.time(12, 0))
        self._entry(self.juan, "MORDIDAS BLANCAS", 1, 7.74, dt.time(7, 0), dt.time(12, 0))
        self._entry(self.pedro, "MORDIDAS AMARILLAS", 1, 10.82, dt.time(7, 0), dt.time(12, 0))
        self._entry(self.juan, "BOTÓN", 5, 20, dt.time(7, 0), dt.time(12, 0))
        self._entry(self.luis, "RECORTE", 6, 20, dt.time(7, 0), dt.time(12, 0))
        self._entry(self.luis, "RECORTE", 8, 20, dt.time(7, 0), dt.time(12, 0))
        self._entry(self.luis, "RECORTE", 3, 20, dt.time(7, 0), dt.time(12, 0))
        self._entry(self.juan, "ANILLAS AMARILLAS", 1, 13.73, dt.time(13, 0), dt.time(18, 0))
        self._entry(self.pedro, "ANILLAS AMARILLAS", 1, 11.91, dt.time(13, 0), dt.time(18, 0))
        self._entry(self.pedro, "RECORTE", 13, 20, dt.time(13, 0), dt.time(18, 0))

    def test_builds_workbook_with_troquelado_sheet(self):
        self._load_sample()
        payload = build_troquelado_xlsx(self.production)
        workbook = load_workbook(BytesIO(payload), data_only=False)
        self.assertIn("CONTROL  DE TROQUELADO", workbook.sheetnames)
        main = workbook["CONTROL  DE TROQUELADO"]
        self.assertEqual(main["A6"].value, dt.time(7, 0))
        self.assertEqual(main["B6"].value, dt.time(12, 0))
        self.assertEqual(main["A11"].value, dt.time(13, 0))
        self.assertEqual(main["B11"].value, dt.time(18, 0))

    def test_notation_matches_the_sample(self):
        self._load_sample()
        payload = build_troquelado_xlsx(self.production)
        workbook = load_workbook(BytesIO(payload), data_only=False)
        main = workbook["CONTROL  DE TROQUELADO"]
        self.assertEqual(main["C6"].value, "4 × 20\n3 × 20")
        self.assertEqual(main["D6"].value, "7.74\nkg")
        self.assertEqual(main["F6"].value, "10.82\nkg")
        self.assertEqual(main["G6"].value, "5 × 20")
        self.assertEqual(main["H6"].value, "6 × 20\n8 × 20\n3 × 20")
        self.assertEqual(main["E11"].value, "13.73\n11.91")
        self.assertEqual(main["H11"].value, "13 × 20")

    def test_subtotal_and_total_formulas(self):
        self._load_sample()
        payload = build_troquelado_xlsx(self.production)
        workbook = load_workbook(BytesIO(payload), data_only=False)
        main = workbook["CONTROL  DE TROQUELADO"]
        self.assertEqual(main["C10"].value, "=4*20+3*20")
        self.assertEqual(main["D10"].value, "=7.74")
        self.assertEqual(main["F10"].value, "=10.82")
        self.assertEqual(main["G10"].value, "=5*20")
        self.assertEqual(main["H10"].value, "=6*20+8*20+3*20")
        self.assertEqual(main["E17"].value, "=13.73+11.91")
        self.assertEqual(main["H17"].value, "=13*20")
        self.assertEqual(main["I6"].value, "=C10+D10+E10+F10+G10+H10")
        self.assertEqual(
            main["G29"].value,
            "=C10+D10+E10+F10+G10+H10+C17+D17+E17+F17+G17+H17",
        )

    def test_error_when_no_entries(self):
        with self.assertRaises(TroqueladoReportError):
            build_troquelado_xlsx(self.production)

    def test_error_when_more_than_two_franjas(self):
        self._entry(self.juan, "BOTÓN", 5, 20, dt.time(7, 0), dt.time(12, 0))
        self._entry(self.juan, "BOTÓN", 5, 20, dt.time(13, 0), dt.time(18, 0))
        self._entry(self.juan, "BOTÓN", 5, 20, dt.time(19, 0), dt.time(23, 0))
        with self.assertRaises(TroqueladoReportError):
            build_troquelado_xlsx(self.production)

    def test_download_view_requires_login_and_returns_xlsx(self):
        self._load_sample()
        reverse_name = reverse("productions:troquelado_report_xlsx", args=[self.production.pk])
        response = self.client.get(reverse_name)
        self.assertNotEqual(response.status_code, 200)

        self.client.force_login(self.user)
        response = self.client.get(reverse_name)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("CONTROL_TROQUELADO", response["Content-Disposition"])

        workbook = load_workbook(BytesIO(response.content), data_only=False)
        main = workbook["CONTROL  DE TROQUELADO"]
        self.assertEqual(main["C10"].value, "=4*20+3*20")

    def test_prepare_troquelado_xlsx_for_pdf(self):
        self._load_sample()
        payload = build_troquelado_xlsx(self.production)
        pdf_ready = load_workbook(
            BytesIO(_prepare_troquelado_xlsx_for_pdf(payload)),
            data_only=False,
        )
        tareos = [s for s in pdf_ready.worksheets if s.title.startswith("TAREO TROQUELADO · ")]
        self.assertEqual(len(tareos), 1)
        tareo = tareos[0]
        self.assertNotEqual(tareo.sheet_state, "hidden")
        self.assertTrue(tareo.print_area.endswith("$A$1:$P$50"))
        self.assertEqual(tareo.page_setup.orientation, "portrait")
        self.assertTrue(tareo.sheet_properties.pageSetUpPr.fitToPage)
        self.assertEqual(tareo.page_setup.fitToWidth, 1)
        self.assertEqual(tareo.page_setup.fitToHeight, 1)
        control = pdf_ready["CONTROL  DE TROQUELADO"]
        self.assertIsNotNone(control.print_area)
        self.assertTrue(control.sheet_properties.pageSetUpPr.fitToPage)
        self.assertEqual(control.page_setup.fitToWidth, 1)
        self.assertEqual(control.page_setup.orientation, "landscape")
        for sheet in pdf_ready.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.font:
                        self.assertNotEqual(cell.font.name, "Arial Narrow")

    def test_tareo_sheet_is_filled_with_workers_and_total(self):
        self._load_sample()
        payload = build_troquelado_xlsx(self.production)
        workbook = load_workbook(BytesIO(payload), data_only=False)
        self.assertIn("TAREO TROQUELADO · CHARLES", workbook.sheetnames)
        self.assertNotIn("EMPAQUE", workbook.sheetnames)
        tareo = workbook["TAREO TROQUELADO · CHARLES"]
        self.assertEqual(tareo["E2"].value, "TAREO DE PERSONAL \nÁREA DE TROQUELADO")
        self.assertEqual(tareo["O2"].value, "SPM-FOT-001")
        self.assertEqual(tareo["O4"].value, "AGOSTO 2026")
        self.assertEqual(tareo["D7"].value, self.production.customer.name)
        self.assertEqual(tareo["L7"].value, self.production.main_product.description)
        self.assertEqual(tareo["D8"].value, '=+TEXT(F8,"dddd")')
        self.assertEqual(tareo["F8"].value.date(), self.production.reception_date)
        self.assertEqual(tareo["L8"].value, "DÍA")
        self.assertEqual(tareo["D9"].value, dt.time(7, 0))
        self.assertEqual(tareo["L9"].value, dt.time(18, 0))
        self.assertEqual(tareo["D10"].value, "DANIELA RIVAS")
        self.assertEqual(tareo["L10"].value, "CHARLES")
        self.assertEqual(tareo["B10"].value, "SUPERVISOR 1:")
        self.assertEqual(tareo["J10"].value, "CUADRILLA")
        self.assertEqual(tareo["B12"].value, "N°")
        self.assertEqual(tareo["C12"].value, "APELLIDOS Y NOMBRES")
        self.assertEqual(tareo["K12"].value, "MENÚ / CENA")
        self.assertEqual(tareo["M12"].value, "TOTAL (PESO)")
        self.assertEqual(tareo["O12"].value, "Importe s/.")
        rows = {}
        for row in range(13, 16):
            name = tareo.cell(row=row, column=3).value
            if name:
                rows[name] = tareo.cell(row=row, column=13).value
        self.assertEqual(
            rows,
            {
                "JUAN PEREZ": 201.47,
                "LUIS ROJAS": 400.0,
                "PEDRO TORRES": 282.73,
            },
        )
        self.assertEqual(tareo["K33"].value, "TOTAL")
        self.assertEqual(tareo["M33"].value, "=SUM(M13:N32)")
        self.assertEqual(tareo["B34"].value, "OBSERVACIONES:")
        self.assertEqual(tareo["B36"].value, "* Especificar solo si son varios PRODUCTOS.")
        self.assertEqual(tareo["B37"].value, "RESUMEN:")
        self.assertEqual(tareo["C37"].value, "N° de PLACA / PRODUCTO *")
        self.assertEqual(tareo["G37"].value, "TOTAL RECIBIDO (KG)")
        self.assertEqual(tareo["K37"].value, "P.U. (S/.)")
        self.assertEqual(tareo["N37"].value, "IMPORTE TOTAL (S/.)")
        self.assertEqual(tareo["B38"].value, 1)
        self.assertEqual(tareo["B39"].value, 2)
        self.assertEqual(tareo["B40"].value, 3)
        self.assertEqual(tareo["C38"].value, "07:00 – 12:00")
        self.assertEqual(tareo["G38"].value, 598.56)
        self.assertEqual(tareo["C39"].value, "13:00 – 18:00")
        self.assertEqual(tareo["G39"].value, 285.64)
        self.assertEqual(tareo["B41"].value, "TOTAL PROCESADO")
        self.assertEqual(tareo["G41"].value, 884.2)
        self.assertEqual(tareo["C46"].value, "Responsable de Planilla")
        self.assertEqual(tareo["J46"].value, "V°B° Gerencia General")
        self.assertEqual(tareo["B48"].value, "RESPONSABLE DEL REGISTRO")
        self.assertEqual(tareo["B49"].value, "NOMBRE:")
        self.assertEqual(tareo["D49"].value, "=+D10")
        self.assertEqual(tareo["K49"].value, "FECHA:")
        self.assertEqual(tareo["M49"].value, "=+F8")
        self.assertEqual(tareo["B50"].value, "CARGO")
        self.assertEqual(tareo["D50"].value, "SUPERVISOR DE PRODUCCIÓN")
        self.assertEqual(tareo["K50"].value, "FIRMA:")
        self.assertEqual(tareo.print_area, "")
        self.assertEqual(tareo.page_setup.scale, 85)
        self.assertIsNone(tareo.page_setup.fitToWidth)
        self.assertEqual(tareo.page_setup.orientation, "portrait")
        self.assertEqual(tareo.row_dimensions[35].height, 24)
        self.assertEqual(tareo.row_dimensions[42].height, 21)
        self.assertEqual(tareo.row_dimensions[51].height, 12.75)
        self.assertEqual(tareo.column_dimensions["A"].width, 1.42578125)
        self.assertEqual(tareo.column_dimensions["K"].width, 12.28515625)
        self.assertEqual(tareo["B12"].fill.fgColor.rgb, "00D9E2F3")
        self.assertEqual(tareo["B37"].fill.fgColor.rgb, "00B4C6E7")
        self.assertEqual(tareo["M33"].fill.fgColor.rgb, "FFFFFF00")
        self.assertEqual(tareo["C38"].fill.fgColor.rgb, "00D5A6BD")
        self.assertEqual(tareo["G38"].fill.fgColor.rgb, "FFFFFF00")
        self.assertEqual(tareo["B38"].fill.fgColor.rgb, "00F2F2F2")

    def test_tareo_allows_up_to_twenty_workers_per_page(self):
        for number in range(19):
            worker = Worker.objects.create(
                internal_code=f"TROQ-X{number}",
                full_name=f"TRABAJADOR {number}",
                active=True,
                crew=self.crew,
            )
            self._entry(worker, "BOTÓN", 1, 1, dt.time(7, 0), dt.time(12, 0))
        payload = build_troquelado_xlsx(self.production)
        workbook = load_workbook(BytesIO(payload), data_only=False)
        tareo = workbook["TAREO TROQUELADO · CHARLES"]
        worker_names = [
            tareo.cell(row=row, column=3).value
            for row in range(13, 33)
        ]
        self.assertEqual(len([name for name in worker_names if name]), 19)
        self.assertEqual(tareo["K33"].value, "TOTAL")
        self.assertEqual(tareo["M33"].value, "=SUM(M13:N32)")

    def test_tareo_rejects_more_than_twenty_workers_per_page(self):
        for number in range(21):
            worker = Worker.objects.create(
                internal_code=f"TROQ-Y{number}",
                full_name=f"TRABAJADOR {number}",
                active=True,
                crew=self.crew,
            )
            self._entry(worker, "BOTÓN", 1, 1, dt.time(7, 0), dt.time(12, 0))
        with self.assertRaises(TroqueladoReportError):
            build_troquelado_xlsx(self.production)

    def test_one_tareo_page_per_crew(self):
        other_crew = Crew.objects.create(code="TROQ-02", name="LAS ESTRELLAS")
        rosa = Worker.objects.create(
            internal_code="TROQ-W7",
            full_name="ROSA FLORES",
            active=True,
            crew=other_crew,
        )
        self._entry(self.juan, "BOTÓN", 5, 20, dt.time(7, 0), dt.time(12, 0))
        self._entry(rosa, "BOTÓN", 3, 20, dt.time(7, 0), dt.time(12, 0), crew=other_crew)
        payload = build_troquelado_xlsx(self.production)
        workbook = load_workbook(BytesIO(payload), data_only=False)
        tareos = [s for s in workbook.sheetnames if s.startswith("TAREO TROQUELADO · ")]
        self.assertEqual(
            tareos,
            ["TAREO TROQUELADO · CHARLES", "TAREO TROQUELADO · LAS ESTRELLAS"],
        )
        charles = workbook[tareos[0]]
        self.assertEqual(charles["L10"].value, "CHARLES")
        self.assertEqual(charles["M13"].value, 100.0)
        self.assertEqual(charles["G38"].value, 100.0)
        self.assertEqual(charles["G41"].value, 100.0)
        estrellas = workbook[tareos[1]]
        self.assertEqual(estrellas["L10"].value, "LAS ESTRELLAS")
        self.assertEqual(estrellas["M13"].value, 60.0)
        self.assertEqual(estrellas["G38"].value, 60.0)
        self.assertEqual(estrellas["G41"].value, 60.0)

    def test_download_page_shows_download_buttons(self):
        self.client.force_login(self.user)
        page = self.client.get(reverse("productions:troquelado_create", args=[self.production.pk]))
        self.assertContains(page, "Descargar control de troquelado Excel")
        self.assertContains(page, "Descargar control de troquelado PDF")
        self.assertContains(page, "Agregar un trabajador y su cuadrilla")

    def test_quick_create_worker_and_crew(self):
        self.client.force_login(self.user)
        url = reverse("productions:troquelado_worker_quick_create", args=[self.production.pk])
        response = self.client.post(url, {"name": "   rosa   flores  ", "crew": "las estrellas"})
        self.assertEqual(response.status_code, 302)
        worker = Worker.objects.get(internal_code__startswith="TROQ-W", full_name="ROSA FLORES")
        self.assertTrue(worker.active)
        self.assertEqual(worker.position, "Troquelador")
        crew = Crew.objects.get(code__startswith="TROQ-", name="LAS ESTRELLAS")
        self.assertTrue(crew.active)
        self.assertEqual(worker.crew_id, crew.pk)

    def test_quick_create_reuses_existing_worker_and_crew(self):
        self.client.force_login(self.user)
        url = reverse("productions:troquelado_worker_quick_create", args=[self.production.pk])
        worker_count = Worker.objects.count()
        crew_count = Crew.objects.count()
        response = self.client.post(url, {"name": "  juan perez ", "crew": "charles"})
        self.assertEqual(response.status_code, 302)
        self.assertEqual(Worker.objects.count(), worker_count)
        self.assertEqual(Crew.objects.count(), crew_count)
        self.assertEqual(Worker.objects.filter(full_name="JUAN PEREZ").count(), 1)

    def test_quick_create_ignores_workers_from_other_areas(self):
        nuq_crew = Crew.objects.create(code="NUQ-01", name="LAS ESTRELLAS")
        Crew.objects.create(code="NUQ-02", name="CRISTHIAN")
        Worker.objects.create(
            internal_code="NUQ-W1",
            full_name="JOSE GARCIA",
            active=True,
            crew=nuq_crew,
        )
        self.client.force_login(self.user)
        url = reverse("productions:troquelado_worker_quick_create", args=[self.production.pk])
        response = self.client.post(url, {"name": "jose garcia", "crew": "cristhian"})
        self.assertEqual(response.status_code, 302)
        troq_worker = Worker.objects.get(
            internal_code__startswith="TROQ-W", full_name="JOSE GARCIA"
        )
        self.assertTrue(troq_worker.internal_code.startswith("TROQ-W"))
        troq_crew = Crew.objects.get(code__startswith="TROQ-", name="CRISTHIAN")
        self.assertEqual(troq_worker.crew_id, troq_crew.pk)
        self.assertEqual(
            Worker.objects.get(internal_code="NUQ-W1").crew_id,
            nuq_crew.pk,
        )
        self.assertEqual(Worker.objects.filter(full_name="JOSE GARCIA").count(), 2)

    def test_quick_create_allows_crew_name_used_in_another_area(self):
        Crew.objects.create(code="NUQ-01", name="CRISTHIAN")
        self.client.force_login(self.user)
        url = reverse("productions:troquelado_worker_quick_create", args=[self.production.pk])
        response = self.client.post(url, {"name": "jose", "crew": "cristhian"})
        self.assertEqual(response.status_code, 302)
        self.assertEqual(Crew.objects.filter(name="CRISTHIAN").count(), 2)
        troq_crew = Crew.objects.get(code__startswith="TROQ-", name="CRISTHIAN")
        self.assertFalse(troq_crew.code.startswith("NUQ-"))
        worker = Worker.objects.get(internal_code__startswith="TROQ-W", full_name="JOSE")
        self.assertEqual(worker.crew_id, troq_crew.pk)

    def test_quick_create_reuses_troq_worker_without_moving_crew(self):
        other_crew = Crew.objects.create(code="TROQ-99", name="OTRA CUADRILLA")
        carlos = Worker.objects.create(
            internal_code="TROQ-W4",
            full_name="CARLOS LOPEZ",
            active=True,
            crew=other_crew,
        )
        self.client.force_login(self.user)
        url = reverse("productions:troquelado_worker_quick_create", args=[self.production.pk])
        worker_count = Worker.objects.count()
        response = self.client.post(url, {"name": "carlos lopez", "crew": "nueva cuadrilla"})
        self.assertEqual(response.status_code, 302)
        self.assertEqual(Worker.objects.count(), worker_count)
        carlos.refresh_from_db()
        self.assertEqual(carlos.crew_id, other_crew.pk)
        crew = Crew.objects.get(code__startswith="TROQ-", name="NUEVA CUADRILLA")
        self.assertNotEqual(crew.pk, other_crew.pk)
        page = self.client.get(
            reverse("productions:troquelado_create", args=[self.production.pk])
        )
        self.assertContains(page, "ya existía en OTRA CUADRILLA")

    def test_quick_created_worker_appears_in_form(self):
        self.client.force_login(self.user)
        url = reverse("productions:troquelado_worker_quick_create", args=[self.production.pk])
        self.client.post(url, {"name": "ROSA FLORES", "crew": "LAS ESTRELLAS"})
        page = self.client.get(reverse("productions:troquelado_create", args=[self.production.pk]))
        self.assertContains(page, "ROSA FLORES")
        self.assertContains(page, "LAS ESTRELLAS")

    def test_quick_create_requires_login(self):
        url = reverse("productions:troquelado_worker_quick_create", args=[self.production.pk])
        response = self.client.post(url, {"name": "ROSA FLORES", "crew": "LAS ESTRELLAS"})
        self.assertNotEqual(response.status_code, 200)

    def test_quick_create_blocked_when_approved(self):
        self.production.status = ProductionOrder.Status.APPROVED
        self.production.save(update_fields=["status"])
        self.client.force_login(self.user)
        url = reverse("productions:troquelado_worker_quick_create", args=[self.production.pk])
        response = self.client.post(url, {"name": "ROSA FLORES", "crew": "LAS ESTRELLAS"})
        self.assertEqual(response.status_code, 403)

    @patch("productions.views.build_troquelado_pdf", return_value=b"%PDF-1.4\n%%EOF")
    def test_pdf_download_uses_the_troquelado_endpoint(self, mocked_builder):
        self.client.force_login(self.user)
        response = self.client.get(
            reverse("productions:troquelado_report_pdf", args=[self.production.pk])
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/pdf")
        self.assertTrue(response.content.startswith(b"%PDF"))
        self.assertIn("CONTROL_TROQUELADO_PP_300", response["Content-Disposition"])

    def test_quick_capture_creates_entry_via_json(self):
        self.client.force_login(self.user)
        url = reverse("productions:troquelado_quick_capture", args=[self.production.pk])
        response = self.client.post(
            url,
            {
                "worker": self.juan.pk,
                "product_type": "BOTÓN",
                "cajas": "5",
                "kg_por_caja": "20",
                "shift": "DAY",
                "start_time": "07:00",
                "end_time": "12:00",
            },
        )
        self.assertEqual(response.status_code, 200)
        payload = json.loads(response.content)
        self.assertTrue(payload["ok"])
        self.assertTrue(payload["entry_id"])
        entry = TroqueladoEntry.objects.get(pk=payload["entry_id"])
        self.assertEqual(entry.worker_id, self.juan.pk)
        self.assertEqual(entry.crew_id, self.crew.pk)
        self.assertEqual(entry.product_type, "BOTÓN")
        self.assertEqual(entry.cajas, 5)
        self.assertEqual(entry.weight_kg, 100)
        self.assertEqual(payload["worker_kg_display"], "100.00")
        self.assertEqual(payload["grand_total_display"], "100.00")
        self.assertEqual(payload["crew_kg_display"], "100.00")
        self.assertEqual(payload["record_count"], 1)
        self.assertEqual(payload["cajas_total"], 5)
        self.assertEqual(payload["crew_name"], "CHARLES")
        self.assertEqual(payload["worker_name"], "JUAN PEREZ")
        card = payload["record_card"]
        self.assertEqual(card["entry_id"], entry.pk)
        self.assertIn("JUAN PEREZ", card["title"])
        self.assertIn("CHARLES", card["title"])
        self.assertIn("5 cajas", card["detail"])
        self.assertTrue(card["edit_url"].startswith(reverse("productions:operational_entry_update", args=[self.production.pk, "troquelado", entry.pk])))
        self.assertTrue(card["delete_url"].startswith(reverse("productions:operational_entry_delete", args=[self.production.pk, "troquelado", entry.pk])))

    def test_quick_capture_updates_stats_for_existing_entries(self):
        self._entry(self.juan, "ANILLAS BLANCAS", 4, 20, dt.time(7, 0), dt.time(12, 0))
        self.client.force_login(self.user)
        url = reverse("productions:troquelado_quick_capture", args=[self.production.pk])
        response = self.client.post(
            url,
            {
                "worker": self.juan.pk,
                "product_type": "BOTÓN",
                "cajas": "5",
                "kg_por_caja": "20",
                "shift": "DAY",
                "start_time": "07:00",
                "end_time": "12:00",
            },
        )
        payload = json.loads(response.content)
        self.assertEqual(payload["worker_kg_display"], "180.00")
        self.assertEqual(payload["grand_total_display"], "180.00")
        self.assertEqual(payload["percent"], 100)
        self.assertEqual(payload["record_count"], 2)
        self.assertEqual(payload["cajas_total"], 9)
        self.assertTrue(payload["categories"])
        boton = next(c for c in payload["categories"] if c["label"] == "Botón")
        self.assertEqual(boton["kg_display"], "100.00")

    def test_quick_capture_requires_product_type(self):
        self.client.force_login(self.user)
        url = reverse("productions:troquelado_quick_capture", args=[self.production.pk])
        response = self.client.post(
            url,
            {"worker": self.juan.pk, "cajas": "5", "kg_por_caja": "20"},
        )
        self.assertEqual(response.status_code, 400)
        payload = json.loads(response.content)
        self.assertFalse(payload["ok"])
        self.assertIn("product_type", payload["errors"])

    def test_quick_capture_requires_cajas(self):
        self.client.force_login(self.user)
        url = reverse("productions:troquelado_quick_capture", args=[self.production.pk])
        response = self.client.post(
            url,
            {"worker": self.juan.pk, "product_type": "BOTÓN", "kg_por_caja": "20"},
        )
        self.assertEqual(response.status_code, 400)
        self.assertFalse(json.loads(response.content)["ok"])

    def test_quick_capture_rejects_unknown_or_foreign_worker(self):
        self.client.force_login(self.user)
        url = reverse("productions:troquelado_quick_capture", args=[self.production.pk])
        nuq = Worker.objects.create(
            internal_code="NUQ-W1",
            full_name="OTRA AREA",
            active=True,
        )
        response = self.client.post(
            url,
            {"worker": nuq.pk, "product_type": "BOTÓN", "cajas": "5", "kg_por_caja": "20"},
        )
        self.assertEqual(response.status_code, 400)
        self.assertFalse(json.loads(response.content)["ok"])
        response = self.client.post(
            url,
            {"worker": 999999, "product_type": "BOTÓN", "cajas": "5", "kg_por_caja": "20"},
        )
        self.assertEqual(response.status_code, 400)

    def test_quick_capture_requires_worker_crew(self):
        no_crew = Worker.objects.create(
            internal_code="TROQ-W9",
            full_name="SIN CUADRILLA",
            active=True,
        )
        self.client.force_login(self.user)
        url = reverse("productions:troquelado_quick_capture", args=[self.production.pk])
        response = self.client.post(
            url,
            {"worker": no_crew.pk, "product_type": "BOTÓN", "cajas": "5", "kg_por_caja": "20"},
        )
        self.assertEqual(response.status_code, 400)
        self.assertFalse(json.loads(response.content)["ok"])

    def test_quick_capture_requires_login(self):
        url = reverse("productions:troquelado_quick_capture", args=[self.production.pk])
        response = self.client.post(url, {"worker": self.juan.pk, "cajas": "1", "kg_por_caja": "1"})
        self.assertNotEqual(response.status_code, 200)

    def test_quick_capture_blocked_when_approved(self):
        self.production.status = ProductionOrder.Status.APPROVED
        self.production.save(update_fields=["status"])
        self.client.force_login(self.user)
        url = reverse("productions:troquelado_quick_capture", args=[self.production.pk])
        response = self.client.post(url, {"worker": self.juan.pk, "cajas": "1", "kg_por_caja": "1"})
        self.assertEqual(response.status_code, 403)

    def test_quick_capture_uses_the_workers_own_crew(self):
        other_crew = Crew.objects.create(code="TROQ-99", name="OTRA CUADRILLA")
        other = Worker.objects.create(
            internal_code="TROQ-W8",
            full_name="DE OTRA CUADRILLA",
            active=True,
            crew=other_crew,
        )
        self.client.force_login(self.user)
        url = reverse("productions:troquelado_quick_capture", args=[self.production.pk])
        response = self.client.post(
            url,
            {
                "worker": other.pk,
                "crew": self.crew.pk,
                "product_type": "BOTÓN",
                "cajas": "5",
                "kg_por_caja": "20",
                "shift": "DAY",
                "start_time": "07:00",
                "end_time": "12:00",
            },
        )
        self.assertEqual(response.status_code, 200)
        entry = TroqueladoEntry.objects.get(production=self.production, worker=other)
        self.assertEqual(entry.crew_id, other_crew.pk)
        self.assertFalse(
            TroqueladoEntry.objects.filter(production=self.production, crew=self.crew).exists()
        )

    def test_quick_capture_page_has_quick_context(self):
        self.client.force_login(self.user)
        page = self.client.get(
            reverse("productions:troquelado_create", args=[self.production.pk])
            + f"?crew={self.crew.pk}&worker={self.juan.pk}"
        )
        self.assertContains(page, "CAPTURA RÁPIDA")
        self.assertContains(page, "troquelado-quick-data")
        workers = page.context["troquelado_quick_workers"]
        self.assertEqual(len(workers), 3)
        juan = next(w for w in workers if w["pk"] == self.juan.pk)
        self.assertFalse(juan["has_entries"])
        self.assertEqual(juan["shift"], "DAY")
        self.assertEqual(juan["start_time"], "06:00")

    def test_quick_capture_page_reflects_existing_entries(self):
        self._entry(self.juan, "ANILLAS BLANCAS", 4, 20, dt.time(7, 0), dt.time(12, 0))
        self.client.force_login(self.user)
        page = self.client.get(
            reverse("productions:troquelado_create", args=[self.production.pk])
            + f"?crew={self.crew.pk}"
        )
        workers = page.context["troquelado_quick_workers"]
        juan = next(w for w in workers if w["pk"] == self.juan.pk)
        self.assertTrue(juan["has_entries"])
        self.assertEqual(juan["kg_display"], "80.00")
        self.assertEqual(juan["product_type"], "ANILLAS BLANCAS")
        self.assertEqual(juan["start_time"], "07:00")
