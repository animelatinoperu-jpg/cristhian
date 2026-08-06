import datetime as dt
from io import BytesIO
from unittest.mock import patch

from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from django.urls import reverse
from openpyxl import load_workbook

from productions.models import (
    Crew,
    Customer,
    PlateCrewEntry,
    PlateEntry,
    PlatePackagingEntry,
    PlatePosition,
    Product,
    ProductionOrder,
    ReceptionEntry,
    TemplateVersion,
    Tunnel,
    TunnelCrewEntry,
    TunnelEntry,
    TunnelFill,
    TunnelPackagingEntry,
    TunnelRack,
    User,
    Vehicle,
)
from productions.services.reception_tareo_report import build_reception_tareo_xlsx
from productions.services.reception_tareo_report_pdf import _prepare_tareo_xlsx_for_pdf


class ReceptionTareoReportTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user("manager", password="Secure-test-123")
        self.user.is_superuser = True
        self.user.is_staff = True
        self.user.save(update_fields=["is_superuser", "is_staff"])
        customer = Customer.objects.create(name="GLOBAL TOP FOOD PERÚ S.A.C.")
        product = Product.objects.create(code="RM-001", description="POTA ENTERA")
        template = TemplateVersion.objects.create(
            code="PP-V1",
            file=SimpleUploadedFile("t.xlsm", b"x"),
            original_filename="t.xlsm",
            sha256="d" * 64,
            uploaded_by=self.user,
        )
        self.production = ProductionOrder.objects.create(
            number=300,
            plant_lot="1064PPF03082026",
            customer_lot="PPF03082026",
            customer=customer,
            process="POTA A GRANEL",
            main_product=product,
            reception_date=dt.date(2026, 8, 3),
            production_date=dt.date(2026, 8, 3),
            shift=ProductionOrder.Shift.DAY,
            template_version=template,
            created_by=self.user,
        )
        crew = Crew.objects.create(code="RM-CUAD-LUIS", name="LUIS")
        vehicle = Vehicle.objects.create(plate="ABC-123")
        ReceptionEntry.objects.create(
            production=self.production,
            responsible=self.user,
            observation="",
            date=self.production.reception_date,
            vehicle=vehicle,
            car_number="1",
            product=product,
            crew=crew,
            container="1",
            weight_kg="542.60",
            time=dt.time(7, 0),
        )

    def test_uses_the_official_fileteros_template_and_preserves_formulas(self):
        payload = build_reception_tareo_xlsx(self.production)
        workbook = load_workbook(BytesIO(payload), data_only=False)

        self.assertEqual(workbook.sheetnames, ["POTA ENTERA", "CUADRILLA 1", "CUADRILLA 2"])
        reception = workbook["POTA ENTERA"]
        self.assertEqual(reception["E2"].value, "REGISTRO DE PESOS POR CUADRILLA")
        self.assertEqual(reception["B7"].font.name, "Arial")
        self.assertEqual(reception["B7"].font.sz, 13)
        self.assertEqual(reception["B7"].alignment.horizontal, "centerContinuous")
        self.assertEqual(reception["E10"].value, "GLOBAL TOP FOOD PERÚ S.A.C.")
        self.assertEqual(reception["T10"].value, "PPF03082026")
        self.assertEqual(reception["C14"].value, "POTA ENTERA")
        self.assertEqual(reception["C15"].value, "ABC-123")
        self.assertEqual(reception["C17"].value, "LUIS")
        self.assertEqual(reception["C18"].value, 542.6)
        for coordinate in ("C14", "C15", "C16", "C17", "C18"):
            self.assertEqual(reception[coordinate].font.name, "Arial")
            self.assertEqual(reception[coordinate].font.sz, 13)
            self.assertEqual(reception[coordinate].alignment.horizontal, "center")
            self.assertEqual(reception[coordinate].alignment.vertical, "center")
        self.assertEqual(reception["C75"].value, "=SUM(C18:C74)")
        self.assertNotIn("CODIGOS", workbook.defined_names)
        self.assertNotIn("LISTA", workbook.defined_names)
        self.assertEqual(len(workbook._external_links), 0)

        tareo = workbook["CUADRILLA 1"]
        self.assertEqual(tareo["E2"].value, "TAREO DE PERSONAL \nÁREA DE FILETEROS")
        self.assertEqual(tareo["B29"].value, "PESO DE LIMPIEZA DE CONO DE POTA POR CUADRILLA")
        self.assertEqual(tareo["C16"].font.name, "Arial")
        self.assertEqual(tareo["C16"].font.sz, 13)
        self.assertEqual(tareo["C17"].font.name, "Arial")
        self.assertEqual(tareo["C17"].font.sz, 13)
        self.assertEqual(tareo["K17"].alignment.horizontal, "center")
        self.assertEqual(tareo["C16"].value, "=+'POTA ENTERA'!D81")
        self.assertEqual(tareo["M16"].value, "=+'POTA ENTERA'!F81")
        self.assertEqual(tareo["M32"].value, 0)
        self.assertGreaterEqual(len(tareo._images), 1)

        pdf_ready = load_workbook(
            BytesIO(_prepare_tareo_xlsx_for_pdf(payload)),
            data_only=False,
        )
        self.assertEqual(str(pdf_ready["POTA ENTERA"].print_area), "'POTA ENTERA'!$B$2:$V$80")
        self.assertEqual(str(pdf_ready["CUADRILLA 1"].print_area), "'CUADRILLA 1'!$B$2:$P$47")
        self.assertEqual(str(pdf_ready["CUADRILLA 2"].print_area), "'CUADRILLA 2'!$B$2:$P$47")

    def test_download_is_a_separate_tareo_button_and_valid_xlsx(self):
        self.client.force_login(self.user)
        page = self.client.get(reverse("productions:reception_create", args=[self.production.pk]))
        self.assertContains(page, "Descargar tareo Excel")
        self.assertContains(page, "Descargar tareo PDF")
        self.assertNotContains(page, "Descargar reporte Excel")
        self.assertNotContains(page, "Descargar reporte PDF")

        response = self.client.get(reverse("productions:reception_tareo_xlsx", args=[self.production.pk]))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertTrue(response.content.startswith(b"PK"))
        self.assertIn("FILETEROS-POTA_TAREO_PP_300", response["Content-Disposition"])

    @patch("productions.views.build_reception_tareo_pdf", return_value=b"%PDF-1.4\n%%EOF")
    def test_pdf_download_uses_the_tareo_endpoint(self, mocked_builder):
        self.client.force_login(self.user)
        response = self.client.get(
            reverse("productions:reception_tareo_pdf", args=[self.production.pk])
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/pdf")
        self.assertTrue(response.content.startswith(b"%PDF"))
        self.assertIn("FILETEROS-POTA_TAREO_PP_300", response["Content-Disposition"])
        mocked_builder.assert_called_once_with(self.production)

    def test_cone_weight_sums_tunnel_and_plate_trays_and_split_50_50(self):
        cone_product = Product.objects.create(
            code="PP-013", description="CONOS DE POTA", presentation=""
        )
        crew_one = Crew.objects.create(code="C1", name="CUADRILLA 1")
        crew_two = Crew.objects.create(code="C2", name="CUADRILLA 2")

        # Las dos cuadrillas aparecen en la recepción del tareo.
        vehicle = Vehicle.objects.create(plate="XYZ-999")
        for car_number, crew in (("1", crew_one), ("2", crew_two)):
            ReceptionEntry.objects.create(
                production=self.production,
                responsible=self.user,
                observation="",
                date=self.production.reception_date,
                vehicle=vehicle,
                car_number=car_number,
                product=self.production.main_product,
                crew=crew,
                container="1",
                weight_kg="100.00",
                time=dt.time(7, 0),
            )

        # tray_kg = 10: 6 bandejas en túnel + 4 bandejas en plaquero (bachada)
        # = 10 bandejas x 10 kg = 100 kg -> 50 kg por cuadrilla.
        self.production.template_version.rules = {"tray_kg": 10}
        self.production.template_version.save()

        tunnel = Tunnel.objects.create(code="T1", name="Túnel 1")
        fill = TunnelFill.objects.create(
            production=self.production,
            tunnel=tunnel,
            fill_number=1,
            date=self.production.production_date,
            supervisor=self.user,
        )
        rack = TunnelRack.objects.create(
            fill=fill, code="R01", position_key="T1!R01", max_trays=50
        )
        common = {"production": self.production, "responsible": self.user, "observation": ""}
        TunnelEntry.objects.create(
            **common,
            rack=rack,
            product=cone_product,
            tray_count=6,
            date=self.production.production_date,
        )

        position = PlatePosition.objects.create(
            template_version=self.production.template_version,
            plate_rack=PlatePosition.PlateRack.P1,
            position_key="ENV. PLACAS!E5",
            display_name="P1 · posición 1",
        )
        PlateEntry.objects.create(
            **common,
            date=self.production.production_date,
            shift=self.production.shift,
            position=position,
            product=cone_product,
            tray_count=4,
        )

        payload = build_reception_tareo_xlsx(self.production)
        workbook = load_workbook(BytesIO(payload), data_only=False)
        # 100 kg total -> 50 kg por cuadrilla.
        self.assertEqual(workbook["CUADRILLA 1"]["M32"].value, 50.0)
        self.assertEqual(workbook["CUADRILLA 2"]["M32"].value, 50.0)

    def test_cone_weight_ignores_packaging_and_crew_trays(self):
        cone_product = Product.objects.create(
            code="PP-013", description="CONOS DE POTA", presentation=""
        )
        crew_one = Crew.objects.create(code="C1", name="CUADRILLA 1")
        crew_two = Crew.objects.create(code="C2", name="CUADRILLA 2")

        vehicle = Vehicle.objects.create(plate="XYZ-999")
        for car_number, crew in (("1", crew_one), ("2", crew_two)):
            ReceptionEntry.objects.create(
                production=self.production,
                responsible=self.user,
                observation="",
                date=self.production.reception_date,
                vehicle=vehicle,
                car_number=car_number,
                product=self.production.main_product,
                crew=crew,
                container="1",
                weight_kg="100.00",
                time=dt.time(7, 0),
            )

        self.production.template_version.rules = {"tray_kg": 10}
        self.production.template_version.save()

        tunnel = Tunnel.objects.create(code="T1", name="Túnel 1")
        fill = TunnelFill.objects.create(
            production=self.production,
            tunnel=tunnel,
            fill_number=1,
            date=self.production.production_date,
            supervisor=self.user,
        )
        rack = TunnelRack.objects.create(
            fill=fill, code="R01", position_key="T1!R01", max_trays=50
        )
        common = {"production": self.production, "responsible": self.user, "observation": ""}
        position = PlatePosition.objects.create(
            template_version=self.production.template_version,
            plate_rack=PlatePosition.PlateRack.P1,
            position_key="ENV. PLACAS!E5",
            display_name="P1 · posición 1",
        )

        # Bultos de envasado y bandejas por cuadrilla NO cuentan.
        TunnelPackagingEntry.objects.create(
            production=self.production,
            responsible=self.user,
            observation="",
            date=self.production.reception_date,
            pallet_number=1,
            product=cone_product,
            package_count=10,
        )
        TunnelCrewEntry.objects.create(
            **common,
            fill=fill,
            rack=rack,
            product=cone_product,
            crew=crew_one,
            page_or_block="PAGINA 1",
            tray_count=30,
            date=self.production.production_date,
        )

        payload = build_reception_tareo_xlsx(self.production)
        workbook = load_workbook(BytesIO(payload), data_only=False)
        # Sin bandejas en túneles/plaqueros -> 0 kg por cuadrilla.
        self.assertEqual(workbook["CUADRILLA 1"]["M32"].value, 0.0)
        self.assertEqual(workbook["CUADRILLA 2"]["M32"].value, 0.0)
