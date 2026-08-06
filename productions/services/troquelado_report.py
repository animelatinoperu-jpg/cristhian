from __future__ import annotations

import datetime as dt
from collections import OrderedDict
from decimal import Decimal
from io import BytesIO
from pathlib import Path

from copy import copy

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from productions.models import TroqueladoEntry

TEMPLATE_PATH = (
    Path(__file__).resolve().parents[1]
    / "report_templates"
    / "CONTROL_TROQUELADO.xlsx"
)

SHEET_NAME = "CONTROL  DE TROQUELADO"
TOTAL_GENERAL_CELL = "G29"

# Hoja de tareo: una página por cuadrilla, formato oficial del tareo de
# personal (estilo SPM-FPO-009) con el área de TROQUELADO.
TAREO_SHEET_PREFIX = "TAREO TROQUELADO · "
TAREO_SOURCE_SHEET = "EMPAQUE"
TAREO_REGISTRO_CODE = "SPM-FOT-001"
TAREO_MAX_WORKERS = 20  # filas 13-32 de la página
TAREO_LAST_ROW = 50

# Estilo del tareo oficial: idéntico al tareo de nucas (SPM-FPO-009).
TAREO_HEADER_FILL = "D9E2F3"        # encabezados de la tabla
TAREO_RESUMEN_FILL = "B4C6E7"       # RESUMEN / TOTAL PROCESADO / RESPONSABLE
TAREO_RESUMEN_NUM_FILL = "F2F2F2"   # numeración del RESUMEN
TAREO_RESUMEN_LABEL_FILL = "D5A6BD" # valor placa/franja del RESUMEN
TAREO_KG_FILL = "FFFFFF00"          # peso en el RESUMEN y TOTAL
TAREO_SIGN_FILL = "D9E2F3"          # rótulos NOMBRE/CARGO/FIRMA
TAREO_BORDER_COLOR = "000000"

# Columnas de la plantilla para cada categoría.
CATEGORY_COLUMNS = {
    "ANILLAS BLANCAS": "C",
    "MORDIDAS BLANCAS": "D",
    "ANILLAS AMARILLAS": "E",
    "MORDIDAS AMARILLAS": "F",
    "BOTÓN": "G",
    "RECORTE": "H",
}

# Categorías que se registran "N cajas x W kg" (múltiplos de peso).
MULTIPLES_CATEGORIES = {"ANILLAS BLANCAS", "BOTÓN", "RECORTE"}
# Categorías que se registran en kg directos.
KG_CATEGORIES = {"MORDIDAS BLANCAS", "MORDIDAS AMARILLAS", "ANILLAS AMARILLAS"}
# Categorías en kg que llevan la leyenda "kg" como última línea.
KG_WITH_UNIT_CATEGORIES = {"MORDIDAS BLANCAS", "MORDIDAS AMARILLAS"}

# Bloques de la plantilla: cada franja (H.I/H.F) ocupa un bloque.
# - ancla: primera fila del bloque (filas de datos y de tiempo).
# - fila_ultima: última fila de datos de categorías.
# - fila_subtotal: fila de subtotales por categoría.
BLOCKS = (
    {"anchor": 6, "last_category": 9, "subtotal": 10},
    {"anchor": 11, "last_category": 16, "subtotal": 17},
)
MAX_BLOCKS = len(BLOCKS)


class TroqueladoReportError(Exception):
    pass


def _decimal(value):
    return Decimal(str(value or 0)).quantize(Decimal("0.01"))


def _fmt_kg(value):
    """Formato corto de kg: sin ceros a la derecha (20, 20.6, 7.74)."""
    value = _decimal(value)
    text = format(value, "f")
    if "." in text:
        text = text.rstrip("0").rstrip(".")
    return text


def _fmt_factor(value):
    text = _fmt_kg(value)
    if text == "-0":
        text = "0"
    return text


def _notation_entries(entries, category):
    """Lista de líneas de notación de una categoría dentro de una franja."""
    if category in MULTIPLES_CATEGORIES:
        return [
            f"{int(_decimal(entry.cajas))} × {_fmt_kg(entry.kg_por_caja)}"
            for entry in entries
        ]
    if category in KG_CATEGORIES:
        lines = [_fmt_kg(entry.weight_kg) for entry in entries]
        if lines and category in KG_WITH_UNIT_CATEGORIES:
            lines.append("kg")
        return lines
    return []


def _subtotal_formula(entries, category):
    """Fórmula del subtotal igual a la usada en la plantilla oficial."""
    if not entries:
        return None
    if category in MULTIPLES_CATEGORIES:
        parts = [
            f"{_fmt_factor(entry.cajas)}*{_fmt_factor(entry.kg_por_caja)}"
            for entry in entries
        ]
    else:
        parts = [_fmt_factor(entry.weight_kg) for entry in entries]
    return "=" + "+".join(parts)


def _normalize_f_column(ws):
    """La plantilla combina F6:F10 y F11:F17 (dato + subtotal). Las separa para
    que MORDIDAS AMARILLAS tenga su celda de subtotal como las demás columnas,
    copiando el estilo de la celda de subtotal del mismo bloque."""
    for anchor, data_rows in ((6, 4), (11, 6)):
        subtotal = anchor + data_rows
        merged = f"F{anchor}:F{subtotal}"
        if any(str(rng) == merged for rng in ws.merged_cells.ranges):
            ws.unmerge_cells(merged)
            ws.merge_cells(f"F{anchor}:F{subtotal - 1}")
            target = ws[f"F{subtotal}"]
            target._style = copy(ws[f"C{subtotal}"]._style)


def _clear_block(ws, block):
    anchor = block["anchor"]
    subtotal = block["subtotal"]
    for row in range(anchor, subtotal + 1):
        for column in range(1, 11):
            cell = ws.cell(row=row, column=column)
            if isinstance(cell, MergedCell):
                continue
            cell.value = None


def _clear_template(ws):
    _normalize_f_column(ws)
    for block in BLOCKS:
        _clear_block(ws, block)
    for row in range(BLOCKS[1]["subtotal"] + 1, 28):
        for column in range(1, 11):
            cell = ws.cell(row=row, column=column)
            if isinstance(cell, MergedCell):
                continue
            cell.value = None
    ws[TOTAL_GENERAL_CELL] = None


def _franjas(entries):
    """Agrupa los registros por franja (inicio, fin) conservando el orden."""
    by_franja = OrderedDict()
    for entry in entries:
        key = (entry.start_time, entry.end_time)
        by_franja.setdefault(key, []).append(entry)
    franjas = []
    for (start, end), group in by_franja.items():
        group.sort(key=lambda entry: (entry.worker.full_name, entry.pk))
        franjas.append({"start": start, "end": end, "entries": group})
    franjas.sort(key=lambda franja: franja["start"] or dt.time(0))
    return franjas


def _tareo_responsible_name(entries):
    for entry in entries:
        full = entry.responsible.get_full_name().strip()
        if full:
            return full.upper()
        if entry.responsible.username:
            return entry.responsible.username.upper()
    return ""


def _tareo_border(style="thin"):
    side = Side(style=style, color=TAREO_BORDER_COLOR)
    return Border(left=side, right=side, top=side, bottom=side)


def _sheet_safe_name(name):
    for char in "[]:*?/\\":
        name = name.replace(char, "-")
    name = name.strip()
    return (name or "SIN NOMBRE")[:24]


def _spanish_month(value):
    """Devuelve MES YYYY en español (p. ej. 'AGOSTO 2026')."""
    months = {
        1: "ENERO", 2: "FEBRERO", 3: "MARZO", 4: "ABRIL", 5: "MAYO", 6: "JUNIO",
        7: "JULIO", 8: "AGOSTO", 9: "SEPTIEMBRE", 10: "OCTUBRE",
        11: "NOVIEMBRE", 12: "DICIEMBRE",
    }
    return f"{months.get(value.month, '')} {value.year}"


def _tareo_unique_name(wb, base_name):
    name = base_name
    counter = 2
    while name in wb.sheetnames:
        suffix = f" ({counter})"
        name = f"{base_name[:31 - len(suffix)]}{suffix}"
        counter += 1
    return name


def _paint_tareo_cell(ws, ref, value=None, *, size=10, bold=False, fill=None,
                      h="center", v="center", wrap=False, sides="all",
                      fmt=None, color="000000"):
    """Aplica el estilo exacto del tareo de nucas a una celda.

    sides: combinación de 'l','r','t','b' (bordes delgados), "all" (los 4),
    "none" (sin bordes) o "medium" (los 4 medianos).
    """
    cell = ws[ref]
    if value is not None:
        cell.value = value
    cell.font = Font(name="Arial", size=size, bold=bold, color=color)
    cell.alignment = Alignment(horizontal=h, vertical=v, wrap_text=wrap)
    if sides == "all":
        cell.border = _tareo_border("thin")
    elif sides == "medium":
        side = Side(style="medium", color=TAREO_BORDER_COLOR)
        cell.border = Border(left=side, right=side, top=side, bottom=side)
    elif sides == "none":
        cell.border = Border()
    else:
        cell.border = Border(
            left=Side(style="thin", color=TAREO_BORDER_COLOR) if "l" in sides else None,
            right=Side(style="thin", color=TAREO_BORDER_COLOR) if "r" in sides else None,
            top=Side(style="thin", color=TAREO_BORDER_COLOR) if "t" in sides else None,
            bottom=Side(style="thin", color=TAREO_BORDER_COLOR) if "b" in sides else None,
        )
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    if fmt:
        cell.number_format = fmt
    return cell


def _merge_tareo_exact(ws, cell_range, value=None, **kwargs):
    ws.merge_cells(cell_range)
    _paint_tareo_cell(ws, cell_range.split(":")[0], value, **kwargs)


def _build_tareo_sheet(wb, production, crew_name, entries):
    """Construye una página TAREO DE PERSONAL · ÁREA DE TROQUELADO con el
    formato oficial de nucas (SPM-FPO-009), una por cuadrilla."""
    if TAREO_SOURCE_SHEET in wb.sheetnames:
        del wb[TAREO_SOURCE_SHEET]
    base_name = f"{TAREO_SHEET_PREFIX}{_sheet_safe_name(crew_name)}"
    ws = wb.create_sheet(_tareo_unique_name(wb, base_name))

    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.scale = 85
    ws.page_margins.left = 0.28
    ws.page_margins.right = 0.1181102362204725
    ws.page_margins.top = 0.26
    ws.page_margins.bottom = 0.1181102362204725

    # Ancho por defecto de columna igual al del tareo oficial de nucas.
    ws.sheet_format.defaultColWidth = 11.28515625
    ws.sheet_format.baseColWidth = 10

    # Anchos exactos del tareo oficial de nucas. Las columnas C, E, F, G, J,
    # M y P NO usan el ancho por defecto: en la plantilla oficial quedaron
    # definidas con el ancho de la columna vecina (rangos B:C, D:G, I:J,
    # L:M y O:P), por lo que se fijan explícitamente aquí.
    for column, width in {
        "A": 1.42578125, "B": 7.7109375, "C": 7.7109375, "D": 6.7109375,
        "E": 6.7109375, "F": 6.7109375, "G": 6.7109375, "H": 7.140625,
        "I": 6.7109375, "J": 6.7109375, "K": 12.28515625, "L": 6.7109375,
        "M": 6.7109375, "N": 7.0, "O": 6.85546875, "P": 6.85546875,
    }.items():
        ws.column_dimensions[column].width = width

    ws.row_dimensions[1].height = 4.5
    for row in (2, 3, 4, 5):
        ws.row_dimensions[row].height = 18.75
    ws.row_dimensions[6].height = 6.75
    for row in (7, 8, 9, 10):
        ws.row_dimensions[row].height = 24
    ws.row_dimensions[11].height = 10.5
    ws.row_dimensions[12].height = 27.75
    for row in range(13, 33):
        ws.row_dimensions[row].height = 20.25
    for row in (33, 34, 35):
        ws.row_dimensions[row].height = 24
    ws.row_dimensions[36].height = 15.75
    ws.row_dimensions[37].height = 18.75
    for row in (38, 39, 40):
        ws.row_dimensions[row].height = 26.25
    for row in (41, 42):
        ws.row_dimensions[row].height = 21
    for row in range(43, 48):
        ws.row_dimensions[row].height = 12.75
    ws.row_dimensions[48].height = 15
    for row in (49, 50):
        ws.row_dimensions[row].height = 21.75
    for row in range(51, 101):
        ws.row_dimensions[row].height = 12.75

    # Título y bloque de control del documento.
    _merge_tareo_exact(ws, "B2:D5", None, size=10, h="center", v="top",
                       wrap=False, sides="all")
    _merge_tareo_exact(ws, "E2:L5", "TAREO DE PERSONAL \nÁREA DE TROQUELADO",
                       size=18, bold=True, h="center", v="center", wrap=True,
                       sides="all")
    _merge_tareo_exact(ws, "M2:N2", "Registro", size=9, bold=True, h="left",
                       v="center", wrap=False, sides="all")
    _merge_tareo_exact(ws, "M3:N3", "Versión", size=9, bold=True, h="left",
                       v="center", wrap=False, sides="all")
    _merge_tareo_exact(ws, "M4:N4", "Fecha", size=9, bold=True, h="left",
                       v="center", wrap=False, sides="all")
    _merge_tareo_exact(ws, "M5:N5", "Página", size=9, bold=True, h="left",
                       v="center", wrap=False, sides="all")
    _merge_tareo_exact(ws, "O2:P2", TAREO_REGISTRO_CODE, size=9, h="center",
                       v="center", wrap=False, sides="all")
    _merge_tareo_exact(ws, "O3:P3", 1, size=9, h="center", v="center",
                       wrap=False, sides="all")
    _merge_tareo_exact(ws, "O4:P4", _spanish_month(production.reception_date),
                       size=9, h="center", v="center", wrap=False, sides="all",
                       fmt="@")
    _merge_tareo_exact(ws, "O5:P5", 1, size=9, h="center", v="center",
                       wrap=False, sides="all")

    # Datos de cabecera.
    times = [
        value
        for entry in entries
        for value in (entry.start_time, entry.end_time)
        if value
    ]
    supervisor = _tareo_responsible_name(entries)
    customer = production.customer.name if production.customer else ""
    product = (
        production.main_product.description
        if production.main_product and production.main_product.description
        else ""
    )

    # Fila 7: cliente y producto.
    _merge_tareo_exact(ws, "B7:C7", "CLIENTE:", size=10, bold=True,
                       h="center", v="center", wrap=False, sides="none")
    _merge_tareo_exact(ws, "D7:H7", customer, size=10, bold=True,
                       h="center", v="center", wrap=True, sides="b")
    _merge_tareo_exact(ws, "J7:K7", "PRODUCTO:", size=10, bold=True,
                       h="center", v="center", wrap=False, sides="none")
    _merge_tareo_exact(ws, "L7:P7", product, size=10, bold=True,
                       h="center", v="center", wrap=True, sides="b")

    # Fila 8: día / fecha y turno.
    _merge_tareo_exact(ws, "B8:C8", "DÍA / FECHA:", size=10, bold=True,
                       h="center", v="center", wrap=False, sides="none")
    _merge_tareo_exact(ws, "D8:E8", '=+TEXT(F8,"dddd")', size=11, bold=True,
                       h="center", v="center", wrap=True, sides="tb",
                       fmt="mm-dd-yy")
    _merge_tareo_exact(ws, "F8:G8", production.reception_date, size=11,
                       bold=True, h="center", v="center", wrap=True,
                       sides="tb", fmt="dd/mm/yyyy")
    _merge_tareo_exact(ws, "J8:K8", "TURNO:", size=10, bold=True,
                       h="center", v="center", wrap=False, sides="none")
    _merge_tareo_exact(ws, "L8:P8", production.get_shift_display().upper(),
                       size=10, bold=True, h="center", v="center", wrap=True,
                       sides="tb")

    # Fila 9: hora de inicio y término.
    _merge_tareo_exact(ws, "B9:C9", "HORA INICIO:", size=10, bold=True,
                       h="center", v="center", wrap=False, sides="none")
    _merge_tareo_exact(ws, "D9:H9", min(times) if times else None, size=10,
                       bold=True, h="center", v="center", wrap=True,
                       sides="tb", fmt="h:mm")
    _merge_tareo_exact(ws, "J9:K9", "HORA TERMINO:", size=10, bold=True,
                       h="center", v="center", wrap=False, sides="none")
    _merge_tareo_exact(ws, "L9:P9", max(times) if times else None, size=10,
                       bold=True, h="center", v="center", wrap=False,
                       sides="tb", fmt="hh:mm:ss")

    # Fila 10: supervisor y cuadrilla.
    _merge_tareo_exact(ws, "B10:C10", "SUPERVISOR 1:", size=10, bold=True,
                       h="center", v="center", wrap=False, sides="none")
    _merge_tareo_exact(ws, "D10:H10", supervisor, size=10, bold=True,
                       h="center", v="center", wrap=True, sides="tb")
    _merge_tareo_exact(ws, "J10:K10", "CUADRILLA", size=10, bold=True,
                       h="center", v="center", wrap=False, sides="none")
    _merge_tareo_exact(ws, "L10:P10", crew_name, size=10, bold=True,
                       h="center", v="center", wrap=True, sides="tb")

    # Pesos por trabajador.
    by_worker = OrderedDict()
    for entry in entries:
        by_worker.setdefault(entry.worker, Decimal("0"))
        by_worker[entry.worker] += Decimal(str(entry.weight_kg or 0))
    workers = sorted(
        by_worker,
        key=lambda worker: (worker.full_name or "").casefold(),
    )
    if len(workers) > TAREO_MAX_WORKERS:
        raise TroqueladoReportError(
            f"La cuadrilla {crew_name} tiene {len(workers)} trabajadores y el "
            f"tareo oficial solo admite {TAREO_MAX_WORKERS} por página."
        )

    # Encabezado de la tabla.
    _paint_tareo_cell(ws, "B12", "N°", size=11, bold=True,
                      fill=TAREO_HEADER_FILL, h="center", v="center",
                      wrap=True, sides="all")
    _merge_tareo_exact(ws, "C12:J12", "APELLIDOS Y NOMBRES", size=11,
                       bold=True, fill=TAREO_HEADER_FILL, h="center",
                       v="center", wrap=False, sides="all")
    _merge_tareo_exact(ws, "K12:L12", "MENÚ / CENA", size=11, bold=True,
                       fill=TAREO_HEADER_FILL, h="center", v="center",
                       wrap=True, sides="all")
    _merge_tareo_exact(ws, "M12:N12", "TOTAL (PESO)", size=11, bold=True,
                       fill=TAREO_HEADER_FILL, h="center", v="center",
                       wrap=True, sides="all")
    _merge_tareo_exact(ws, "O12:P12", "Importe s/.", size=11, bold=True,
                       fill=TAREO_HEADER_FILL, h="center", v="center",
                       wrap=True, sides="all")

    for index, worker in enumerate(workers, start=1):
        row = 12 + index
        _paint_tareo_cell(ws, f"B{row}", index, size=10, h="center",
                          v="center", wrap=True, sides="all")
        _merge_tareo_exact(ws, f"C{row}:J{row}", worker.full_name, size=10,
                           h="center", v="center", wrap=True, sides="all")
        _merge_tareo_exact(ws, f"K{row}:L{row}", None, size=10, h="center",
                           v="center", wrap=True, sides="all")
        _merge_tareo_exact(ws, f"M{row}:N{row}", float(by_worker[worker]),
                           size=10, h="center", v="center", wrap=True,
                           sides="all")
        _merge_tareo_exact(ws, f"O{row}:P{row}", None, size=10, h="center",
                           v="center", wrap=True, sides="all")

    # Filas de la tabla sin trabajadores: conservan los bordes de la plantilla
    # oficial (el tareo de nucas imprime la tabla completa aunque haya filas
    # vacías), por eso se pintan con sus merges y bordes.
    for row in range(13 + len(workers), 33):
        _paint_tareo_cell(ws, f"B{row}", None, size=10, h="center",
                          v="center", wrap=True, sides="all")
        _merge_tareo_exact(ws, f"C{row}:J{row}", None, size=10, h="center",
                           v="center", wrap=True, sides="all")
        _merge_tareo_exact(ws, f"K{row}:L{row}", None, size=10, h="center",
                           v="center", wrap=True, sides="all")
        _merge_tareo_exact(ws, f"M{row}:N{row}", None, size=10, h="center",
                           v="center", wrap=True, sides="all")
        _merge_tareo_exact(ws, f"O{row}:P{row}", None, size=10, h="center",
                           v="center", wrap=True, sides="all")

    # TOTAL.
    _merge_tareo_exact(ws, "C33:G33", None, size=10, h="center", v="center",
                       wrap=True, sides="none")
    _merge_tareo_exact(ws, "I33:J33", None, size=10, h="center", v="center",
                       wrap=True, sides="none")
    _merge_tareo_exact(ws, "K33:L33", "TOTAL", size=10, bold=True,
                       h="center", v="center", wrap=True, sides="all")
    _merge_tareo_exact(ws, "M33:N33", "=SUM(M13:N32)", size=10, bold=True,
                       fill=TAREO_KG_FILL, h="center", v="center", wrap=True,
                       sides="lrb")
    _merge_tareo_exact(ws, "O33:P33", None, size=10, h="center", v="center",
                       wrap=True, sides="all")

    # Observaciones.
    _paint_tareo_cell(ws, "B34", "OBSERVACIONES:", size=10, bold=True,
                      h="left", v=None, wrap=False, sides="none")
    for column in "DEFGHIJKLMNOP":
        _paint_tareo_cell(ws, f"{column}34", None, size=10, h="center",
                          v="center", wrap=True, sides="b")
    for column in "BCDEFGHIJKLMNOP":
        _paint_tareo_cell(ws, f"{column}35", None, size=10, h="center",
                          v="center", wrap=True, sides="b")

    _merge_tareo_exact(ws, "B36:P36",
                       "* Especificar solo si son varios PRODUCTOS.", size=8,
                       bold=True, h="left", v="center", wrap=False, sides="tb")

    # RESUMEN: una fila por franja con su peso.
    _paint_tareo_cell(ws, "B37", "RESUMEN:", size=9, bold=True,
                      fill=TAREO_RESUMEN_FILL, h=None, v="center",
                      wrap=False, sides="all")
    _merge_tareo_exact(ws, "C37:F37", "N° de PLACA / PRODUCTO *", size=10,
                       bold=True, fill=TAREO_RESUMEN_FILL, h="center",
                       v="center", wrap=False, sides="all")
    _merge_tareo_exact(ws, "G37:J37", "TOTAL RECIBIDO (KG)", size=10,
                       bold=True, fill=TAREO_RESUMEN_FILL, h="center",
                       v="center", wrap=False, sides="all")
    _merge_tareo_exact(ws, "K37:M37", "P.U. (S/.)", size=10, bold=True,
                       fill=TAREO_RESUMEN_FILL, h="center", v="center",
                       wrap=False, sides="all")
    _merge_tareo_exact(ws, "N37:P37", "IMPORTE TOTAL (S/.)", size=10,
                       bold=True, fill=TAREO_RESUMEN_FILL, h="center",
                       v="center", wrap=False, sides="all")

    franjas = _franjas(entries)
    for index in range(1, 4):
        row = 37 + index
        _paint_tareo_cell(ws, f"B{row}", index, size=10, bold=True,
                          fill=TAREO_RESUMEN_NUM_FILL, h="center",
                          v="center", wrap=False, sides="lrb")
        if index <= len(franjas):
            franja = franjas[index - 1]
            start = franja["start"]
            end = franja["end"]
            label = (
                f"{start.strftime('%H:%M')} – {end.strftime('%H:%M')}"
                if start and end
                else "Sin horas"
            )
            subtotal = sum(
                (Decimal(str(entry.weight_kg or 0)) for entry in franja["entries"]),
                Decimal("0"),
            )
            _merge_tareo_exact(ws, f"C{row}:F{row}", label, size=10,
                               bold=True, fill=TAREO_RESUMEN_LABEL_FILL,
                               h="center", v="center", wrap=False,
                               sides="all")
            _merge_tareo_exact(ws, f"G{row}:J{row}", float(subtotal), size=10,
                               bold=True, fill=TAREO_KG_FILL, h="center",
                               v="center", wrap=False, sides="all")
        else:
            _merge_tareo_exact(ws, f"C{row}:F{row}", None, size=10, bold=True,
                               h="center", v="center", wrap=False, sides="all")
            _merge_tareo_exact(ws, f"G{row}:J{row}", None, size=10, bold=True,
                               h="center", v="center", wrap=False, sides="all")
        _merge_tareo_exact(ws, f"K{row}:M{row}", None, size=10, bold=True,
                           h="center", v="center", wrap=False, sides="all")
        _merge_tareo_exact(ws, f"N{row}:P{row}", None, size=10, bold=True,
                           h="center", v="center", wrap=False, sides="all")

    total = sum(by_worker.values(), Decimal("0"))
    _merge_tareo_exact(ws, "B41:F41", "TOTAL PROCESADO", size=11, bold=True,
                       fill=TAREO_RESUMEN_FILL, h="center", v="center",
                       wrap=False, sides="all")
    _merge_tareo_exact(ws, "G41:J41", float(total), size=11, bold=True,
                       fill=TAREO_RESUMEN_FILL, h="center", v="center",
                       wrap=False, sides="all")
    for column in "KLM":
        _paint_tareo_cell(ws, f"{column}41", None, size=11, bold=True,
                          fill=TAREO_RESUMEN_FILL, h=None, v="center",
                          wrap=False, sides="tb")
    _merge_tareo_exact(ws, "N41:P42", None, size=11, bold=True,
                       fill=TAREO_RESUMEN_FILL, h="center", v="center",
                       wrap=False, sides="medium")

    # Firmas.
    _merge_tareo_exact(ws, "C46:F46", "Responsable de Planilla", size=10,
                       bold=True, h="center", v="center", wrap=False, sides="t")
    _merge_tareo_exact(ws, "J46:M46", "V°B° Gerencia General", size=10,
                       bold=True, h="center", v="center", wrap=False, sides="t")
    _merge_tareo_exact(ws, "B48:P48", "RESPONSABLE DEL REGISTRO", size=9,
                       bold=True, fill=TAREO_RESUMEN_FILL, h="center",
                       v="center", wrap=False, sides="all")
    _merge_tareo_exact(ws, "B49:C49", "NOMBRE:", size=9, bold=True,
                       fill=TAREO_SIGN_FILL, h="left", v="center", wrap=False,
                       sides="lrb")
    _merge_tareo_exact(ws, "D49:J49", "=+D10", size=9, bold=True, h="center",
                       v="center", wrap=False, sides="all")
    _merge_tareo_exact(ws, "K49:L49", "FECHA:", size=9, bold=True,
                       fill=TAREO_SIGN_FILL, h="left", v="center", wrap=False,
                       sides="all")
    _merge_tareo_exact(ws, "M49:P49", "=+F8", size=9, bold=True, h="center",
                       v="center", wrap=False, sides="all", fmt="mm-dd-yy")
    _merge_tareo_exact(ws, "B50:C50", "CARGO", size=9, bold=True,
                       fill=TAREO_SIGN_FILL, h="left", v="center", wrap=False,
                       sides="all")
    _merge_tareo_exact(ws, "D50:J50", "SUPERVISOR DE PRODUCCIÓN", size=9,
                       bold=True, h="center", v="center", wrap=False,
                       sides="all")
    _merge_tareo_exact(ws, "K50:L50", "FIRMA:", size=9, bold=True,
                       fill=TAREO_SIGN_FILL, h="left", v="center", wrap=False,
                       sides="ltb")
    _merge_tareo_exact(ws, "M50:P50", None, size=9, bold=True, h="center",
                       v="center", wrap=False, sides="all")

    return ws


def build_troquelado_xlsx(production):
    if not TEMPLATE_PATH.exists():
        raise TroqueladoReportError("No se encontró la plantilla oficial de troquelado.")

    entries = list(
        TroqueladoEntry.objects.filter(production=production, is_active=True)
        .select_related("crew", "worker")
        .order_by("start_time", "pk")
    )
    entries = [entry for entry in entries if entry.product_type]
    if not entries:
        raise TroqueladoReportError(
            "Todavía no hay registros de troquelado para generar el reporte."
        )

    franjas = _franjas(entries)
    if len(franjas) > MAX_BLOCKS:
        raise TroqueladoReportError(
            f"El parte tiene {len(franjas)} franjas (H.I/H.F) y la plantilla "
            f"oficial solo admite {MAX_BLOCKS}. Agrupe las horas para poder generar el reporte."
        )

    wb = load_workbook(TEMPLATE_PATH, keep_links=False)
    ws = wb[SHEET_NAME]
    _clear_template(ws)

    for franja, block in zip(franjas, BLOCKS):
        anchor = block["anchor"]
        subtotal = block["subtotal"]
        if franja["start"]:
            ws[f"A{anchor}"] = franja["start"]
            ws[f"A{anchor}"].number_format = "hh:mm"
        if franja["end"]:
            ws[f"B{anchor}"] = franja["end"]
            ws[f"B{anchor}"].number_format = "hh:mm"

        for category, column in CATEGORY_COLUMNS.items():
            category_entries = [
                entry
                for entry in franja["entries"]
                if entry.product_type == category
            ]
            lines = _notation_entries(category_entries, category)
            if lines:
                ws[f"{column}{anchor}"] = "\n".join(lines)
                ws[f"{column}{anchor}"].alignment = (
                    ws[f"{column}{anchor}"].alignment.copy(
                        wrap_text=True,
                        vertical="top",
                    )
                )
            formula = _subtotal_formula(category_entries, category)
            if formula:
                ws[f"{column}{subtotal}"] = formula

        total_formula = "=" + "+".join(
            f"{column}{subtotal}" for column in "CDEFGH"
        )
        ws[f"I{anchor}"] = total_formula

    grand_parts = []
    for block in BLOCKS:
        grand_parts.extend(f"{column}{block['subtotal']}" for column in "CDEFGH")
    ws[TOTAL_GENERAL_CELL] = "=" + "+".join(grand_parts)

    # Una página de tareo por cuadrilla.
    by_crew = OrderedDict()
    for entry in entries:
        crew_name = entry.crew.name if entry.crew and entry.crew.name else "SIN CUADRILLA"
        by_crew.setdefault(crew_name, []).append(entry)
    for crew_name, crew_entries in by_crew.items():
        _build_tareo_sheet(wb, production, crew_name, crew_entries)

    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"

    payload = BytesIO()
    wb.save(payload)
    return payload.getvalue()
