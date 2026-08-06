from __future__ import annotations

import datetime as dt
import re
from collections import defaultdict
from copy import copy
from decimal import Decimal
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter

from .crew_control import _normalized_crew_name

TEMPLATE_PATH = (
    Path(__file__).resolve().parents[1]
    / "report_templates"
    / "NUQUERAS_TAREO.xlsx"
)

# El bloque estándar (CHARLES) ocupa las filas 6-39 de la hoja NUQUERAS.
BLOCK_HEADER_ROW = 6
BLOCK_NAMES_ROW = 7
BLOCK_DATA_FIRST = 8
BLOCK_DATA_LAST = 37
BLOCK_SUBTOTAL_ROW = 38
BLOCK_TOTAL_ROW = 39
BLOCK_SPAN = BLOCK_TOTAL_ROW - BLOCK_HEADER_ROW + 1  # 34
BLOCK_GAP = 37  # distancia entre inicios de bloques extra consecutivos

# Primeras filas de los bloques extra ya presentes en la plantilla oficial.
EXTRA_BLOCK_STARTS = (76, 113)

# Las tres cuadrillas base de la plantilla oficial. El label y las filas de
# cada cuadrilla se escriben/leen por su nombre; las cuadrillas nuevas usan
# bloques extra en la hoja NUQUERAS.
BASE_SLOTS = [
    {
        "slot": "CHARLES",
        "sheet": "NUQUERAS",
        "header_row": 6,
        "label": "C6",
        "label_prefix": "CUADRILLA ",
        "names_row": 7,
        "first_row": 8,
        "last_row": 37,
        "subtotal_row": 38,
        "total_row": 39,
        "number_col": "A",
        "block": (8, 37),
        "weight_rows": (6, 39),
        "tareo": "TAREO CHARLES",
    },
    {
        "slot": "PAZ",
        "sheet": "NUQUERAS",
        "header_row": 43,
        "label": "C43",
        "label_prefix": "CUADRILLA ",
        "names_row": 44,
        "first_row": 45,
        "last_row": 71,
        "subtotal_row": 72,
        "total_row": 73,
        "number_col": "B",
        "block": (45, 71),
        "weight_rows": (43, 73),
        "gap_rows": (6, 42),
        "tareo": "TAREO PAZ",
    },
    {
        "slot": "OMAR",
        "sheet": " NUQUERAS omar ",
        "header_row": 6,
        "label": "C6",
        "label_prefix": "",
        "names_row": 7,
        "first_row": 8,
        "last_row": 37,
        "subtotal_row": 38,
        "total_row": 39,
        "number_col": "A",
        "block": (8, 37),
        "weight_rows": (6, 39),
        "tareo": "TAREO OMAR",
    },
]
MAX_WORKERS = 14  # columnas C..P
BASE_SLOT_NAMES = tuple(slot["slot"] for slot in BASE_SLOTS)


class NuqueraTareoReportError(Exception):
    pass


def _decimal(value):
    return Decimal(str(value or 0)).quantize(Decimal("0.01"))


def _spanish_day(value):
    names = (
        "Lunes",
        "Martes",
        "Miércoles",
        "Jueves",
        "Viernes",
        "Sábado",
        "Domingo",
    )
    return names[value.weekday()]


def _spanish_month(value):
    names = (
        "ENERO",
        "FEBRERO",
        "MARZO",
        "ABRIL",
        "MAYO",
        "JUNIO",
        "JULIO",
        "AGOSTO",
        "SEPTIEMBRE",
        "OCTUBRE",
        "NOVIEMBRE",
        "DICIEMBRE",
    )
    return f"{names[value.month - 1]} {value.year}"


def _responsible_name(entries):
    for entry in entries:
        name = entry.responsible.get_full_name().strip()
        if name:
            return name.upper()
        if entry.responsible.username:
            return entry.responsible.username.upper()
    return ""


def extra_block_start(extra_index):
    """Fila inicial de un bloque extra (0 = primer bloque extra)."""
    if extra_index < len(EXTRA_BLOCK_STARTS):
        return EXTRA_BLOCK_STARTS[extra_index]
    last = EXTRA_BLOCK_STARTS[-1]
    return last + (extra_index - len(EXTRA_BLOCK_STARTS) + 1) * BLOCK_GAP


def _extra_layout(extra_index, dest):
    """Layout de un bloque extra (0 = primer bloque extra, fila 76)."""
    crew_index = len(BASE_SLOTS) + extra_index
    return {
        "slot": f"EXTRA{extra_index + 1}",
        "sheet": "NUQUERAS",
        "header_row": dest,
        "label": f"C{dest}",
        "label_prefix": "CUADRILLA ",
        "names_row": dest + 1,
        "first_row": dest + 2,
        "last_row": dest + 31,
        "subtotal_row": dest + 32,
        "total_row": dest + 33,
        "number_col": "A",
        "block": (dest + 2, dest + 31),
        "weight_rows": (dest, dest + 33),
        "gap_rows": (6, dest - 1),
        "tareo": f"TAREO CUADRILLA {crew_index + 1}",
    }


def iter_slot_layouts(workbook):
    """Layouts base + los bloques extra presentes en el workbook (detectados
    por sus hojas TAREO), en orden de aparición."""
    layouts = list(BASE_SLOTS)
    extras = {}
    for name in workbook.sheetnames:
        if not name.startswith("TAREO CUADRILLA "):
            continue
        try:
            crew_index = int(name.rsplit(" ", 1)[1]) - 1
        except ValueError:
            continue
        extra_index = crew_index - len(BASE_SLOTS)
        if extra_index >= 0:
            extras[extra_index] = _extra_layout(extra_index, extra_block_start(extra_index))
    for extra_index in sorted(extras):
        layouts.append(extras[extra_index])
    return layouts


def _resolve_slots(crews):
    """Asigna cada cuadrilla de la producción a un bloque.

    Primero intenta el nombre (CHARLES/PAZ/OMAR); las cuadrillas que no
    coinciden ocupan bloques base libres y, si se acaban, bloques extra
    dinámicos (sin límite).
    """
    normalized = [_normalized_crew_name(crew.name) for crew in crews]
    assigned = {}

    for base in BASE_SLOTS:
        for name in normalized:
            if base["slot"] in name and name not in assigned:
                assigned[name] = base["slot"]
                break

    for base in BASE_SLOTS:
        if base["slot"] in assigned.values():
            continue
        for name in normalized:
            if name not in assigned:
                assigned[name] = base["slot"]
                break

    next_extra = 1
    for name in normalized:
        if name not in assigned:
            assigned[name] = f"EXTRA{next_extra}"
            next_extra += 1

    return [assigned[name] for name in normalized]


def _layout_for_slot_key(slot_key):
    if slot_key in BASE_SLOT_NAMES:
        return next(base for base in BASE_SLOTS if base["slot"] == slot_key)
    extra_index = int(slot_key[len("EXTRA"):]) - 1
    return _extra_layout(extra_index, extra_block_start(extra_index))


def _copy_row_style(source_ws, source_row, target_ws, target_row, max_col=16):
    for column in range(1, max_col + 1):
        src = source_ws.cell(row=source_row, column=column)
        if isinstance(src, MergedCell):
            continue
        dst = target_ws.cell(row=target_row, column=column)
        dst._style = copy(src._style)


def _overlaps_merge(worksheet, coordinate):
    from openpyxl.worksheet.cell_range import CellRange

    min_col, min_row, max_col, max_row = CellRange(coordinate).bounds
    for merged in worksheet.merged_cells.ranges:
        if (
            merged.min_col <= max_col
            and min_col <= merged.max_col
            and merged.min_row <= max_row
            and min_row <= merged.max_row
        ):
            return True
    return False


def _clone_weight_block(worksheet, dest):
    """Copia el bloque estándar (filas 6-39) a la fila `dest` con sus estilos,
    alturas, celdas combinadas y fórmulas de SUBTOTAL/TOTAL ajustadas."""
    for offset in range(BLOCK_SPAN):
        src_row = BLOCK_HEADER_ROW + offset
        dst_row = dest + offset
        _copy_row_style(worksheet, src_row, worksheet, dst_row)
        height = worksheet.row_dimensions[src_row].height
        if height is not None:
            worksheet.row_dimensions[dst_row].height = height

    header = dest
    names = dest + 1
    subtotal = dest + (BLOCK_SUBTOTAL_ROW - BLOCK_HEADER_ROW)
    total = dest + (BLOCK_TOTAL_ROW - BLOCK_HEADER_ROW)
    worksheet[f"A{header}"] = "N°"
    worksheet[f"B{header}"] = "CUADRILLA"
    worksheet[f"C{header}"] = "CUADRILLA "
    data_first = names + 1
    data_last = names + (BLOCK_DATA_LAST - BLOCK_NAMES_ROW)
    for column in range(2, 17):
        letter = get_column_letter(column)
        worksheet[f"{letter}{subtotal}"] = f"=SUM({letter}{data_first}:{letter}{data_last})"
    worksheet[f"B{total}"] = f"=SUM(B{subtotal}:P{subtotal})"

    for coordinate in (f"C{header}:P{header}", f"A{header}:A{names}", f"B{total}:P{total}"):
        if not _overlaps_merge(worksheet, coordinate):
            worksheet.merge_cells(coordinate)


def _retarget_tareo_formula(value, names_row, subtotal_row):
    if not isinstance(value, str) or "NUQUERAS!" not in value:
        return value
    value = re.sub(
        r"(NUQUERAS!\$?[A-Z]+)7\b",
        lambda match: f"{match.group(1)}{names_row}",
        value,
    )
    value = re.sub(
        r"(NUQUERAS!\$?[A-Z]+)38\b",
        lambda match: f"{match.group(1)}{subtotal_row}",
        value,
    )
    return value


def _clone_tareo_sheet(workbook, new_name, names_row, subtotal_row):
    """Copia la hoja TAREO CHARLES como plantilla para una cuadrilla nueva,
    reenrutando sus referencias al bloque correspondiente."""
    source = workbook["TAREO CHARLES"]
    target = workbook.create_sheet(new_name)
    for row in range(1, source.max_row + 1):
        height = source.row_dimensions[row].height
        if height is not None:
            target.row_dimensions[row].height = height
        for column in range(1, 17):
            src = source.cell(row=row, column=column)
            if isinstance(src, MergedCell):
                continue
            dst = target.cell(row=row, column=column)
            dst.value = _retarget_tareo_formula(src.value, names_row, subtotal_row)
            dst._style = copy(src._style)
    for merged in source.merged_cells.ranges:
        target.merge_cells(str(merged))
    for letter, dimension in source.column_dimensions.items():
        if dimension.width is not None:
            target.column_dimensions[letter].width = dimension.width
    target.page_setup.orientation = source.page_setup.orientation
    target.page_setup.paperSize = source.page_setup.paperSize
    return target


def _ensure_slots(workbook, slot_keys):
    """Crea los bloques extra y sus hojas TAREO que falten para las cuadrillas."""
    main_sheet = workbook["NUQUERAS"]
    for slot_key in slot_keys:
        if slot_key in BASE_SLOT_NAMES:
            continue
        extra_index = int(slot_key[len("EXTRA"):]) - 1
        dest = extra_block_start(extra_index)
        if main_sheet.cell(row=dest, column=1).value != "N°":
            _clone_weight_block(main_sheet, dest)
        layout = _extra_layout(extra_index, dest)
        if layout["tareo"] not in workbook.sheetnames:
            _clone_tareo_sheet(
                workbook,
                layout["tareo"],
                layout["names_row"],
                layout["subtotal_row"],
            )


def _clear_block(ws, layout):
    number_col = layout["number_col"]
    for row in range(layout["names_row"], layout["last_row"] + 1):
        for column in range(1, 17):
            if row == layout["names_row"] and column == 2 and number_col == "B":
                continue
            cell = ws.cell(row=row, column=column)
            if isinstance(cell, MergedCell):
                continue
            cell.value = None
    for row in range(layout["first_row"], layout["last_row"] + 1):
        ws[f"{number_col}{row}"] = None


def _fill_weight_block(ws, layout, crew_name, entries):
    """Escribe los nombres y pesos de una cuadrilla en su bloque de la plantilla."""
    workers = []
    by_worker = defaultdict(list)
    for entry in entries:
        if entry.worker not in by_worker:
            workers.append(entry.worker)
        by_worker[entry.worker].append(entry)
    if len(workers) > MAX_WORKERS:
        raise NuqueraTareoReportError(
            f"La cuadrilla {crew_name} tiene {len(workers)} trabajadores y la "
            f"plantilla de nuqueras solo admite {MAX_WORKERS} por bloque."
        )

    ws[layout["label"]] = f"{layout['label_prefix']}{crew_name}"

    for index, worker in enumerate(workers):
        column = 3 + index  # columna C en adelante
        ws.cell(row=layout["names_row"], column=column).value = worker.full_name
        if len(by_worker[worker]) > layout["last_row"] - layout["first_row"] + 1:
            raise NuqueraTareoReportError(
                f"El trabajador {worker.full_name} tiene más de "
                f"{layout['last_row'] - layout['first_row'] + 1} pesos y no cabe "
                "en la plantilla de nuqueras."
            )
        for offset, entry in enumerate(by_worker[worker]):
            ws.cell(
                row=layout["first_row"] + offset,
                column=column,
            ).value = float(_decimal(entry.weight_kg))

    number_col = layout["number_col"]
    counter = 1
    for row in range(layout["first_row"], layout["last_row"] + 1):
        row_has_value = any(
            ws.cell(row=row, column=column).value is not None
            for column in range(3, 3 + len(workers))
        )
        if row_has_value:
            ws[f"{number_col}{row}"] = counter
            counter += 1


def _fill_tareo_sheet(ws, production, slot, crew_name, entries):
    by_worker = defaultdict(list)
    workers = []
    for entry in entries:
        if entry.worker not in by_worker:
            workers.append(entry.worker)
        by_worker[entry.worker].append(entry)

    ws["O4"] = _spanish_month(production.reception_date)
    ws["D7"] = production.customer.name
    ws["L7"] = production.main_product.description.upper()
    ws["F8"] = production.reception_date
    ws["F8"].number_format = "dd/mm/yyyy"
    ws["L8"] = production.get_shift_display().upper()

    time_values = [
        value
        for entry in entries
        for value in (entry.start_time, entry.end_time)
        if value
    ]
    start_time = min(time_values) if time_values else None
    end_time = max(time_values) if time_values else None
    ws["L9"] = end_time
    if end_time:
        ws["L9"].number_format = "hh:mm:ss"
        ws["L9"].alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=False,
        )
    ws["D10"] = _responsible_name(entries)
    ws["L10"] = crew_name

    first_row = 13
    for offset, worker in enumerate(workers):
        row = first_row + offset
        total = sum((_decimal(entry.weight_kg) for entry in by_worker[worker]), Decimal("0"))
        ws.cell(row=row, column=2).value = offset + 1
        ws.cell(row=row, column=3).value = worker.full_name
        ws.cell(row=row, column=11).value = None
        ws.cell(row=row, column=13).value = float(total)

    grand_total = sum(
        (sum((_decimal(entry.weight_kg) for entry in by_worker[worker]), Decimal("0"))
         for worker in workers),
        Decimal("0"),
    )
    ws["B38"] = 1
    ws["C38"] = "NUCA"
    ws["G38"] = float(grand_total)


def _clear_tareo_sheet(ws):
    for row in range(13, 33):
        for column in (2, 3, 11, 13):
            cell = ws.cell(row=row, column=column)
            if not isinstance(cell, MergedCell):
                cell.value = None
    for coordinate in ("D7", "L7", "F8", "L8", "L9", "D10", "L10", "B38", "C38", "G38", "O4"):
        cell = ws[coordinate]
        if not isinstance(cell, MergedCell):
            cell.value = None


def build_nuquera_tareo_xlsx(production):
    if not TEMPLATE_PATH.exists():
        raise NuqueraTareoReportError("No se encontró la plantilla oficial de nuqueras.")

    entries = list(
        NuqueraEntry.objects.filter(production=production, is_active=True)
        .select_related("crew", "worker", "responsible")
        .order_by("start_time", "pk")
    )
    if not entries:
        raise NuqueraTareoReportError("Todavía no hay pesos de nuqueras para generar el tareo.")

    by_crew = defaultdict(list)
    crews = []
    for entry in entries:
        crew_name = _normalized_crew_name(entry.crew.name)
        if crew_name not in by_crew:
            crews.append(entry.crew)
        by_crew[crew_name].append(entry)

    slot_keys = _resolve_slots(crews)

    wb = load_workbook(TEMPLATE_PATH, keep_links=False)
    _ensure_slots(wb, slot_keys)
    main_sheet = wb["NUQUERAS"]
    omar_sheet = wb[" NUQUERAS omar "]

    header_date = production.reception_date
    header_start = None
    header_end = None
    for crew_entries in by_crew.values():
        times = [
            value
            for entry in crew_entries
            for value in (entry.start_time, entry.end_time)
            if value
        ]
        if times:
            header_start = min(header_start, times[0]) if header_start else min(times)
            header_end = max(header_end, times[-1]) if header_end else max(times)

    for sheet in (main_sheet, omar_sheet):
        sheet["N1"] = header_date
        sheet["N1"].number_format = "dd/mm/yyyy"
        sheet["N2"] = header_start
        sheet["N3"] = header_end
        sheet["D5"] = _spanish_day(header_date)
        sheet["N5"] = "NUCA"

    used_tareos = []
    for crew, slot_key in zip(crews, slot_keys):
        layout = _layout_for_slot_key(slot_key)
        used_tareos.append(layout["tareo"])
        crew_entries = by_crew[_normalized_crew_name(crew.name)]
        sheet = wb[layout["sheet"]]
        _clear_block(sheet, layout)
        _fill_weight_block(
            sheet,
            layout,
            _normalized_crew_name(crew.name),
            crew_entries,
        )
        tareo = wb[layout["tareo"]]
        _clear_tareo_sheet(tareo)
        _fill_tareo_sheet(
            tareo,
            production,
            layout["slot"],
            _normalized_crew_name(crew.name),
            crew_entries,
        )

    for layout in iter_slot_layouts(wb):
        if layout["tareo"] in used_tareos:
            continue
        sheet = wb[layout["sheet"]]
        _clear_block(sheet, layout)
        if layout["slot"] in BASE_SLOT_NAMES:
            sheet[layout["label"]] = f"{layout['label_prefix']}{layout['slot']}"

    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"

    payload = BytesIO()
    wb.save(payload)
    return payload.getvalue()


from openpyxl.styles import Alignment  # noqa: E402

from productions.models import NuqueraEntry  # noqa: E402
