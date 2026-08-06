from __future__ import annotations

import os
import shutil
import subprocess
import tempfile
from copy import copy
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter

from .nuquera_tareo_report import (
    NuqueraTareoReportError,
    build_nuquera_tareo_xlsx,
    iter_slot_layouts,
)


class NuqueraTareoPdfError(NuqueraTareoReportError):
    pass


TAREO_PRINT_AREA = "A1:P50"
HEADER_ROWS = (1, 5)
MAX_COL = 16


def _block_has_data(worksheet, first_row, last_row, max_col):
    for row in range(first_row, last_row + 1):
        for column in range(3, max_col + 1):
            if worksheet.cell(row=row, column=column).value is not None:
                return True
    return False


def _active_layouts(workbook):
    """Layouts (cuadrillas) que tienen al menos un peso en la plantilla."""
    active = []
    for layout in iter_slot_layouts(workbook):
        source = workbook[layout["sheet"]]
        first_row, last_row = layout["block"]
        if _block_has_data(source, first_row, last_row, MAX_COL):
            active.append(layout)
    return active


def _crew_sheet_label(layout, source_sheet):
    """Nombre legible de la cuadrilla para nombrar la hoja de pesos."""
    value = source_sheet[layout["label"]].value or ""
    prefix = layout["label_prefix"] or ""
    if value.startswith(prefix):
        value = value[len(prefix):]
    value = value.strip().upper()
    return value or layout["slot"]


def _copy_rows(source, target, first_row, last_row):
    for row in range(first_row, last_row + 1):
        height = source.row_dimensions[row].height
        if height is not None:
            target.row_dimensions[row].height = height
        for column in range(1, MAX_COL + 1):
            src_cell = source.cell(row=row, column=column)
            if isinstance(src_cell, MergedCell):
                continue
            dst_cell = target.cell(row=row, column=column)
            dst_cell.value = src_cell.value
            dst_cell._style = copy(src_cell._style)


def _copy_merged_ranges(source, target, allowed_rows):
    for merged in source.merged_cells.ranges:
        if any(
            low <= merged.min_row and merged.max_row <= high
            for low, high in allowed_rows
        ):
            target.merge_cells(str(merged))


def _copy_column_widths(source, target):
    for column in range(1, MAX_COL + 1):
        letter = get_column_letter(column)
        width = source.column_dimensions[letter].width
        if width is not None:
            target.column_dimensions[letter].width = width


def _create_crew_weight_sheet(workbook, sheet_name, layout):
    """Hoja de impresión por cuadrilla: cabecera del documento + bloque de la
    cuadrilla, copiando valores, estilos, alturas y celdas fusionadas."""
    source = workbook[layout["sheet"]]
    target = workbook.create_sheet(sheet_name)
    first, last = layout["weight_rows"]
    _copy_rows(source, target, *HEADER_ROWS)
    _copy_rows(source, target, first, last)
    _copy_column_widths(source, target)
    _copy_merged_ranges(source, target, [HEADER_ROWS, (first, last)])
    gap = layout.get("gap_rows")
    if gap:
        for row in range(gap[0], gap[1] + 1):
            target.row_dimensions[row].hidden = True
    return target


def _fit_crew_name_row(worksheet, names_row):
    """Ajusta la fila de nombres de la cuadrilla para que se imprima completa y
    centrada: los nombres se envuelven dentro de su celda y la fila se
    autoajusta a su altura (igual que la fila de CHARLES en la plantilla, que
    ya se ve bien). Evita que los nombres largos se desborden hacia arriba o
    se corten en el PDF."""
    for column in range(3, MAX_COL + 1):
        cell = worksheet.cell(row=names_row, column=column)
        if cell.value is None:
            continue
        alignment = copy(cell.alignment)
        alignment.wrap_text = True
        alignment.vertical = "center"
        alignment.horizontal = "center"
        cell.alignment = alignment
    worksheet.row_dimensions[names_row].height = None


def _configure_print(worksheet, print_area, *, fit_height):
    worksheet.print_area = print_area
    worksheet.page_setup.orientation = "portrait"
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = fit_height
    worksheet.page_setup.scale = None
    worksheet.sheet_view.showGridLines = False


def _compress_unused_block_rows(worksheet, block):
    """Encoge las filas vacías dentro de un bloque para que sus datos y la fila
    TOTAL queden juntos en la misma página al escalar (ajuste de escala, no se
    reestructura el contenido)."""
    first_row, last_row = block
    for row in range(first_row, last_row + 1):
        has_value = any(
            worksheet.cell(row=row, column=column).value is not None
            for column in range(1, MAX_COL + 1)
        )
        if not has_value:
            worksheet.row_dimensions[row].height = 1


def _replace_narrow_fonts(worksheet):
    """Arial Narrow no existe en los servidores Linux: se sustituye por Arial.

    Liberation Sans (instalado en el contenedor) es métricamente compatible
    con Arial, así el PDF se ve igual en Windows y en Railway.
    """
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.font and cell.font.name == "Arial Narrow":
                cell.font = cell.font.copy(name="Arial")


def _prepare_tareo_xlsx_for_pdf(xlsx_payload: bytes) -> bytes:
    """Prepara el workbook oficial para imprimir un PDF por cuadrilla:
    cada cuadrilla con pesos sale con su cuadro de pesos y su tareo.

    Las hojas de pesos originales y los tareos de cuadrillas sin datos se
    ocultan. LibreOffice imprime las hojas ocultas si conservan print_area,
    por eso también se quita el área de impresión.
    """
    workbook = load_workbook(BytesIO(xlsx_payload), keep_links=False)
    active_layouts = _active_layouts(workbook)

    for sheet_name in ("NUQUERAS", " NUQUERAS omar "):
        sheet = workbook[sheet_name]
        sheet.print_area = None
        sheet.sheet_state = "hidden"

    active_tareos = {layout["tareo"] for layout in active_layouts}
    for layout in iter_slot_layouts(workbook):
        if layout["tareo"] not in active_tareos:
            tareo = workbook[layout["tareo"]]
            tareo.print_area = None
            tareo.sheet_state = "hidden"

    printed = []
    for layout in active_layouts:
        source_sheet = workbook[layout["sheet"]]
        crew_label = _crew_sheet_label(layout, source_sheet)
        tareo_name = layout["tareo"]
        tareo_sheet = workbook[tareo_name]

        pesos_name = f"PESOS {crew_label}"
        pesos_sheet = _create_crew_weight_sheet(workbook, pesos_name, layout)
        _fit_crew_name_row(pesos_sheet, layout["names_row"])
        _configure_print(
            pesos_sheet,
            f"A1:P{layout['total_row']}",
            fit_height=0,
        )
        _replace_narrow_fonts(pesos_sheet)
        _compress_unused_block_rows(pesos_sheet, layout["block"])
        printed.append(pesos_name)

        _configure_print(tareo_sheet, TAREO_PRINT_AREA, fit_height=1)
        _replace_narrow_fonts(tareo_sheet)
        printed.append(tareo_name)

    remaining = [
        name for name in workbook.sheetnames if name not in printed
    ]
    workbook._sheets = [workbook[name] for name in printed + remaining]

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _libreoffice_binary() -> str:
    binary = shutil.which("soffice") or shutil.which("libreoffice")
    if not binary:
        raise NuqueraTareoPdfError(
            "No se encontró LibreOffice para convertir el tareo de nucas a PDF."
        )
    return binary


def _xlsx_to_pdf_payload(xlsx_payload: bytes, *, stem: str) -> bytes:
    binary = _libreoffice_binary()
    with tempfile.TemporaryDirectory(prefix="nuquera-tareo-pdf-") as temp_name:
        temp_dir = Path(temp_name)
        profile_dir = temp_dir / "libreoffice-profile"
        profile_dir.mkdir()
        input_path = temp_dir / f"{stem}.xlsx"
        output_path = temp_dir / f"{stem}.pdf"
        input_path.write_bytes(_prepare_tareo_xlsx_for_pdf(xlsx_payload))

        result = subprocess.run(
            [
                binary,
                "--headless",
                "--nologo",
                "--nodefault",
                "--nofirststartwizard",
                f"-env:UserInstallation={profile_dir.resolve().as_uri()}",
                "--convert-to",
                "pdf:calc_pdf_Export",
                "--outdir",
                str(temp_dir),
                str(input_path),
            ],
            cwd=temp_dir,
            env=os.environ.copy(),
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            timeout=120,
            check=False,
        )
        if result.returncode != 0 or not output_path.is_file():
            stderr = result.stderr.decode("utf-8", errors="replace").strip()
            stdout = result.stdout.decode("utf-8", errors="replace").strip()
            detail = stderr or stdout or "LibreOffice no generó el PDF."
            raise NuqueraTareoPdfError(
                f"No se pudo convertir el tareo de nucas a PDF: {detail[:300]}"
            )

        payload = output_path.read_bytes()
        if not payload.startswith(b"%PDF"):
            raise NuqueraTareoPdfError("El tareo de nucas generado no es un PDF válido.")
        return payload


def build_nuquera_tareo_pdf(production) -> bytes:
    xlsx_payload = build_nuquera_tareo_xlsx(production)
    return _xlsx_to_pdf_payload(
        xlsx_payload,
        stem=f"NUQUERAS_TAREO_PP_{production.number}",
    )
