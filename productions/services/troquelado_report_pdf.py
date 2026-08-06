from __future__ import annotations

import os
import shutil
import subprocess
import tempfile
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from .troquelado_report import (
    SHEET_NAME,
    TAREO_LAST_ROW,
    TAREO_SHEET_PREFIX,
    TroqueladoReportError,
    build_troquelado_xlsx,
)


class TroqueladoReportPdfError(TroqueladoReportError):
    pass


PRINT_AREA = "A1:J29"


def _configure_print(worksheet):
    worksheet.print_area = PRINT_AREA
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 1
    worksheet.page_setup.scale = None
    worksheet.sheet_view.showGridLines = False


def _replace_narrow_fonts(worksheet):
    """Arial Narrow no existe en los servidores Linux: se sustituye por Arial."""
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.font and cell.font.name == "Arial Narrow":
                cell.font = cell.font.copy(name="Arial")


def _configure_tareo_print(worksheet):
    """Ajuste de impresión del tareo, igual que el de nucas (A1:P50, 1 hoja)."""
    worksheet.print_area = f"A1:P{TAREO_LAST_ROW}"
    worksheet.page_setup.orientation = "portrait"
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 1
    worksheet.page_setup.scale = None
    worksheet.sheet_view.showGridLines = False


def _prepare_troquelado_xlsx_for_pdf(xlsx_payload: bytes) -> bytes:
    """Deja la hoja de control y las páginas de tareo listas para imprimir."""
    workbook = load_workbook(BytesIO(xlsx_payload), keep_links=False)
    control = workbook[SHEET_NAME]
    _configure_print(control)
    _replace_narrow_fonts(control)
    for sheet in workbook.worksheets:
        if sheet.title.startswith(TAREO_SHEET_PREFIX):
            _configure_tareo_print(sheet)
            _replace_narrow_fonts(sheet)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _libreoffice_binary() -> str:
    binary = shutil.which("soffice") or shutil.which("libreoffice")
    if not binary:
        raise TroqueladoReportPdfError(
            "No se encontró LibreOffice para convertir el reporte de troquelado a PDF."
        )
    return binary


def _xlsx_to_pdf_payload(xlsx_payload: bytes, *, stem: str) -> bytes:
    binary = _libreoffice_binary()
    with tempfile.TemporaryDirectory(prefix="troquelado-report-pdf-") as temp_name:
        temp_dir = Path(temp_name)
        profile_dir = temp_dir / "libreoffice-profile"
        profile_dir.mkdir()
        input_path = temp_dir / f"{stem}.xlsx"
        output_path = temp_dir / f"{stem}.pdf"
        input_path.write_bytes(_prepare_troquelado_xlsx_for_pdf(xlsx_payload))

        result = subprocess.run(
            [
                binary,
                "--headless",
                "--nologo",
                "--nodefault",
                "--nofirststartwizard",
                f"-env:UserInstallation={profile_dir.resolve().as_uri()}",
                "--convert-to",
                "pdf:calc_pdf_Export",
                "--outdir",
                str(temp_dir),
                str(input_path),
            ],
            cwd=temp_dir,
            env=os.environ.copy(),
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            timeout=120,
            check=False,
        )
        if result.returncode != 0 or not output_path.is_file():
            stderr = result.stderr.decode("utf-8", errors="replace").strip()
            stdout = result.stdout.decode("utf-8", errors="replace").strip()
            detail = stderr or stdout or "LibreOffice no generó el PDF."
            raise TroqueladoReportPdfError(
                f"No se pudo convertir el reporte de troquelado a PDF: {detail[:300]}"
            )

        payload = output_path.read_bytes()
        if not payload.startswith(b"%PDF"):
            raise TroqueladoReportPdfError(
                "El reporte de troquelado generado no es un PDF válido."
            )
        return payload


def build_troquelado_pdf(production) -> bytes:
    xlsx_payload = build_troquelado_xlsx(production)
    return _xlsx_to_pdf_payload(
        xlsx_payload,
        stem=f"CONTROL_TROQUELADO_PP_{production.number}",
    )
