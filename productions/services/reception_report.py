from __future__ import annotations

from collections import defaultdict
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from productions.models import ReceptionEntry
from .crew_control import _normalized_crew_name


FOREST = "124B3B"
MINT = "DCEBE5"
PAPER = "F3F5F4"
INK = "17211E"
WHITE = "FFFFFF"
LIGHT = "EAF3EF"
SOFT_BLUE = "E6F1F7"


def _decimal_kg(value) -> Decimal:
    return Decimal(str(value or 0)).quantize(Decimal("0.01"))


def _cone_pota_product(product) -> bool:
    haystack = " ".join(
        part
        for part in (
            getattr(product, "code", ""),
            getattr(product, "description", ""),
            getattr(product, "presentation", ""),
        )
        if part
    )
    normalized = _normalized_crew_name(haystack)
    return "CONO" in normalized and "POTA" in normalized


def _style_cell(cell, *, bold=False, fill=None, color=INK, center=False):
    cell.font = Font(name="Arial", size=10, bold=bold, color=color)
    cell.alignment = Alignment(horizontal="center" if center else "left", vertical="center", wrap_text=True)
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    cell.border = Border(
        left=Side(style="thin", color="C9D7D1"),
        right=Side(style="thin", color="C9D7D1"),
        top=Side(style="thin", color="C9D7D1"),
        bottom=Side(style="thin", color="C9D7D1"),
    )


def _build_rows(production):
    entries = (
        ReceptionEntry.objects.filter(production=production, is_active=True)
        .select_related("vehicle", "product", "crew", "responsible")
        .order_by("crew__name", "vehicle__plate", "car_number", "container", "pk")
    )

    detail_rows = []
    crew_rows = {}
    cone_rows = {}
    total_weight = Decimal("0.00")

    for entry in entries:
        crew_name = entry.crew.name if entry.crew else "SIN CUADRILLA"
        normalized_crew = _normalized_crew_name(crew_name) or "SIN CUADRILLA"
        weight = _decimal_kg(entry.weight_kg)
        is_cone_pota = _cone_pota_product(entry.product)

        detail_rows.append(
            {
                "date": entry.date,
                "vehicle": entry.vehicle.plate,
                "car_number": (entry.car_number or "").strip() or "—",
                "product": entry.product.description,
                "crew": normalized_crew,
                "container": (entry.container or "").strip() or "—",
                "weight_kg": weight,
                "is_cone_pota": is_cone_pota,
            }
        )

        crew_row = crew_rows.setdefault(
            normalized_crew,
            {
                "crew": normalized_crew,
                "count": 0,
                "weight": Decimal("0.00"),
                "cone_count": 0,
                "cone_weight": Decimal("0.00"),
            },
        )
        crew_row["count"] += 1
        crew_row["weight"] += weight

        if is_cone_pota:
            crew_row["cone_count"] += 1
            crew_row["cone_weight"] += weight
            cone_row = cone_rows.setdefault(
                normalized_crew,
                {
                    "crew": normalized_crew,
                    "count": 0,
                    "weight": Decimal("0.00"),
                },
            )
            cone_row["count"] += 1
            cone_row["weight"] += weight

        total_weight += weight

    crew_summary = sorted(
        crew_rows.values(),
        key=lambda item: (-item["weight"], item["crew"].casefold()),
    )
    cone_summary = sorted(
        cone_rows.values(),
        key=lambda item: (-item["weight"], item["crew"].casefold()),
    )
    return detail_rows, crew_summary, cone_summary, total_weight


def _write_title_and_meta(ws, production, total_weight):
    ws.merge_cells("A1:G1")
    ws["A1"] = f"TAREO DE RECEPCIÓN · PP {production.number}"
    ws["A1"].font = Font(name="Arial Black", size=14, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=FOREST)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    _style_cell(ws["A1"], bold=True, fill=FOREST, color=WHITE, center=True)
    ws.row_dimensions[1].height = 24

    meta = [
        ("Lote de planta", production.plant_lot),
        ("Lote cliente", production.customer_lot),
        ("Cliente", production.customer.name),
        ("Proceso", production.process),
        ("Fecha recepción", production.reception_date.strftime("%d/%m/%Y")),
        ("Peso total", f"{total_weight:,.2f} kg"),
    ]
    for idx, (label, value) in enumerate(meta, start=3):
        ws[f"A{idx}"] = label
        ws[f"B{idx}"] = value
        _style_cell(ws[f"A{idx}"], bold=True, fill=LIGHT)
        _style_cell(ws[f"B{idx}"], fill="FFFFFF")


def _write_crew_summary(ws, start_row, crew_summary):
    ws[f"D{start_row}"] = "Resumen de cuadrillas"
    _style_cell(ws[f"D{start_row}"], bold=True, fill=MINT)
    ws[f"D{start_row}"].font = Font(name="Arial Black", size=11, bold=True, color=INK)

    headers = ["Cuadrilla", "Registros", "Peso kg", "Cono pota kg"]
    for offset, header in enumerate(headers, start=0):
        cell = ws.cell(row=start_row + 1, column=4 + offset, value=header)
        _style_cell(cell, bold=True, fill=FOREST, color=WHITE, center=True)
        cell.font = Font(name="Arial Black", size=10, bold=True, color=WHITE)

    if crew_summary:
        for row_index, item in enumerate(crew_summary, start=start_row + 2):
            values = [
                item["crew"],
                item["count"],
                f'{item["weight"]:,.2f}',
                f'{item["cone_weight"]:,.2f}',
            ]
            for col, value in enumerate(values, start=4):
                cell = ws.cell(row=row_index, column=col, value=value)
                _style_cell(cell, center=col != 4)
                if row_index % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor=PAPER)
    else:
        ws.merge_cells(start_row=start_row + 2, start_column=4, end_row=start_row + 2, end_column=7)
        ws.cell(row=start_row + 2, column=4, value="Sin registros de cuadrillas")
        _style_cell(ws.cell(row=start_row + 2, column=4), center=True)

    return start_row + 2 + max(len(crew_summary), 1)


def _write_cone_summary(ws, start_row, cone_summary):
    ws[f"D{start_row}"] = "Peso de limpieza de cono de pota por cuadrilla"
    _style_cell(ws[f"D{start_row}"], bold=True, fill=SOFT_BLUE)
    ws[f"D{start_row}"].font = Font(name="Arial Black", size=11, bold=True, color=INK)

    headers = ["Cuadrilla", "Registros", "Peso kg"]
    for offset, header in enumerate(headers, start=0):
        cell = ws.cell(row=start_row + 1, column=4 + offset, value=header)
        _style_cell(cell, bold=True, fill=FOREST, color=WHITE, center=True)
        cell.font = Font(name="Arial Black", size=10, bold=True, color=WHITE)

    if cone_summary:
        for row_index, item in enumerate(cone_summary, start=start_row + 2):
            values = [item["crew"], item["count"], f'{item["weight"]:,.2f}']
            for col, value in enumerate(values, start=4):
                cell = ws.cell(row=row_index, column=col, value=value)
                _style_cell(cell, center=col != 4)
                if row_index % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor=PAPER)
    else:
        ws.merge_cells(start_row=start_row + 2, start_column=4, end_row=start_row + 2, end_column=6)
        ws.cell(row=start_row + 2, column=4, value="Sin registros de CONOS DE POTA")
        _style_cell(ws.cell(row=start_row + 2, column=4), center=True)

    return start_row + 2 + max(len(cone_summary), 1)


def _write_detail_table(ws, start_row, detail_rows):
    headers = ["Fecha", "Vehículo", "Carro", "Producto", "Cuadrilla", "Dino", "Peso kg"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col, value=header)
        _style_cell(cell, bold=True, fill=FOREST, color=WHITE, center=True)
        cell.font = Font(name="Arial Black", size=10, bold=True, color=WHITE)

    for row_index, item in enumerate(detail_rows, start=start_row + 1):
        values = [
            item["date"].strftime("%d/%m/%Y"),
            item["vehicle"],
            item["car_number"],
            item["product"],
            item["crew"],
            item["container"],
            f'{item["weight_kg"]:,.2f}',
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row_index, column=col, value=value)
            _style_cell(cell, center=col != 4)
            if row_index % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=PAPER)


def build_reception_report_xlsx(production):
    detail_rows, crew_summary, cone_summary, total_weight = _build_rows(production)

    wb = Workbook()
    ws = wb.active
    ws.title = "Tareo"
    ws.freeze_panes = "A7"
    ws.sheet_view.showGridLines = False

    _write_title_and_meta(ws, production, total_weight)
    detail_start = _write_crew_summary(ws, 3, crew_summary)
    detail_start = _write_cone_summary(ws, detail_start + 1, cone_summary)
    detail_start += 2
    _write_detail_table(ws, detail_start, detail_rows)

    widths = {1: 14, 2: 18, 3: 12, 4: 34, 5: 18, 6: 10, 7: 12}
    for col_idx, width in widths.items():
        ws.column_dimensions[chr(64 + col_idx)].width = width

    payload = BytesIO()
    wb.save(payload)
    return payload.getvalue()


TAREO_BLUE = "8EAADB"
TAREO_HEADER = "D9E2F3"
TAREO_FOOTER = "B4C6E7"
TAREO_GREY = "E7E6E6"
TAREO_BORDER = "000000"


def _tareo_border(style="thin"):
    side = Side(style=style, color=TAREO_BORDER)
    return Border(left=side, right=side, top=side, bottom=side)


def _tareo_cell(cell, value=None, *, fill=None, bold=False, size=10, center=True, border="thin"):
    if value is not None:
        cell.value = value
    cell.font = Font(name="Arial", size=size, bold=bold, color="000000")
    cell.alignment = Alignment(horizontal="center" if center else "left", vertical="center", wrap_text=True)
    cell.border = _tareo_border(border) if border else Border()
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)


def _merge_tareo(ws, cell_range, value, **kwargs):
    ws.merge_cells(cell_range)
    _tareo_cell(ws[cell_range.split(":")[0]], value, **kwargs)


def _setup_tareo_page(ws):
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.35
    ws.page_margins.bottom = 0.35
    for column, width in {
        "A": 2.2, "B": 7.5, "C": 8.5, "D": 7.5, "E": 7.5, "F": 7.5,
        "G": 7.5, "H": 7.5, "I": 7.5, "J": 7.5, "K": 8.5, "L": 7.5,
        "M": 8.5, "N": 7.5, "O": 8.5, "P": 7.5,
    }.items():
        ws.column_dimensions[column].width = width


def _build_reception_control_sheet(ws, production, detail_rows, total_weight):
    _setup_tareo_page(ws)
    ws.merge_cells("B2:P4")
    _tareo_cell(ws["B2"], "REGISTRO DE PESOS POR CUADRILLA", bold=True, size=18, border="thin")
    ws.row_dimensions[2].height = 24
    for row in (3, 4):
        ws.row_dimensions[row].height = 20

    _merge_tareo(ws, "B6:P6", "DATOS DEL EMPLEADOR", fill=TAREO_BLUE, bold=True, size=12, border="medium")
    _merge_tareo(ws, "B7:D7", "CLIENTE:", bold=True, center=False)
    _merge_tareo(ws, "E7:P7", production.customer.name, center=False)
    _merge_tareo(ws, "B8:D8", "PROCESO:", bold=True, center=False)
    _merge_tareo(ws, "E8:P8", production.process, center=False)
    _merge_tareo(ws, "B9:D9", "LOTE DE PLANTA:", bold=True, center=False)
    _merge_tareo(ws, "E9:H9", production.plant_lot, center=False)
    _merge_tareo(ws, "I9:K9", "FECHA:", bold=True, center=False)
    _merge_tareo(ws, "L9:P9", production.reception_date.strftime("%d/%m/%Y"), center=False)
    _merge_tareo(ws, "B10:D10", "PESO TOTAL:", bold=True, center=False)
    _merge_tareo(ws, "E10:H10", float(total_weight), bold=True)
    ws["E10"].number_format = '#,##0.00 "kg"'

    headers = ["N°", "PLACA", "N° CARRO", "CUADRILLA", "N° DINO", "PRODUCTO", "PESO (kg)"]
    ranges = ["B12:C12", "D12:E12", "F12:G12", "H12:I12", "J12:K12", "L12:N12", "O12:P12"]
    for cell_range, header in zip(ranges, headers):
        _merge_tareo(ws, cell_range, header, fill=TAREO_HEADER, bold=True, size=10)

    for row_no, item in enumerate(detail_rows, start=13):
        values = [
            row_no - 12,
            item["vehicle"],
            item["car_number"],
            item["crew"],
            item["container"],
            item["product"],
            float(item["weight_kg"]),
        ]
        for cell_range, value in zip(ranges, values):
            _merge_tareo(ws, cell_range.replace("12", str(row_no)), value, size=9, center=cell_range not in {"H12:I12", "L12:N12"})
        ws[f"O{row_no}"].number_format = '#,##0.00'
        ws.row_dimensions[row_no].height = 25

    total_row = max(13, 12 + len(detail_rows) + 1)
    _merge_tareo(ws, f"B{total_row}:N{total_row}", "TOTAL", fill=TAREO_BLUE, bold=True, size=11, border="medium")
    _merge_tareo(ws, f"O{total_row}:P{total_row}", float(total_weight), fill=TAREO_GREY, bold=True, size=12, border="medium")
    ws[f"O{total_row}"].number_format = '#,##0.00 "kg"'
    ws.print_area = f"B2:P{total_row}"


def _build_crew_tareo_sheet(ws, production, crew, index):
    _setup_tareo_page(ws)
    crew_name = crew["crew"]
    reception_weight = _decimal_kg(crew["weight"])
    cone_weight = _decimal_kg(crew["cone_weight"])
    # The original tareo uses 40% of fileteo as the washing line.
    washing_weight = (reception_weight * Decimal("0.40")).quantize(Decimal("0.01"))
    total_general = reception_weight + washing_weight + cone_weight

    _merge_tareo(ws, "E2:L5", "TAREO DE PERSONAL\nÁREA DE FILETEROS", bold=True, size=18, border="thin")
    _merge_tareo(ws, "M2:N2", "Registro", bold=True, size=9)
    _merge_tareo(ws, "O2:P2", "SPM-FOT-007", bold=True, size=9)
    _merge_tareo(ws, "M3:N3", "Versión", bold=True, size=9)
    _merge_tareo(ws, "O3:P3", 1, bold=True, size=9)
    _merge_tareo(ws, "M4:N4", "Fecha", bold=True, size=9)
    _merge_tareo(ws, "O4:P4", production.reception_date.strftime("%B %Y").upper(), bold=True, size=9)
    _merge_tareo(ws, "M5:N5", "Página", bold=True, size=9)
    _merge_tareo(ws, "O5:P5", index, bold=True, size=9)

    meta = [
        ("B7:C7", "CLIENTE:", "D7:H7", production.customer.name),
        ("J7:L7", "PRODUCTO:", "M7:P7", production.main_product.description),
        ("B8:C8", "DÍA / FECHA:", "D8:H8", production.reception_date.strftime("%A %d/%m/%Y").upper()),
        ("J8:L8", "TURNO:", "M8:P8", production.get_shift_display().upper()),
        ("B9:D9", "HORA INICIO:", "E9:H9", "-"),
        ("J9:K9", "HORA TERMINO:", "L9:P9", "-"),
        ("B10:C10", "SUPERVISOR:", "D10:H10", "-"),
        ("J10:K10", "PLACA CARRO:", "L10:P10", "VARIOS"),
    ]
    for label_range, label, value_range, value in meta:
        _merge_tareo(ws, label_range, label, bold=True, size=10, center=False)
        _merge_tareo(ws, value_range, value, bold=True, size=10)

    _write_tareo_section(ws, 13, "PESO DE FILETEO POR CUADRILLA", crew_name, reception_weight, crew["count"])
    _write_tareo_section(ws, 21, "PESO DE LAVADO DE FILETE DE POTA POR CUADRILLA", crew_name, washing_weight, crew["count"])
    _write_tareo_section(ws, 29, "PESO DE LIMPIEZA DE CONO DE POTA POR CUADRILLA", crew_name, cone_weight, crew["cone_count"])

    _merge_tareo(ws, "K36:M36", "TOTAL GENERAL", fill=TAREO_BLUE, bold=True, size=10, border="medium")
    _merge_tareo(ws, "N36:P36", float(total_general), fill=TAREO_GREY, bold=True, size=18, border="medium")
    ws["N36"].number_format = '#,##0.00 "kg"'
    _merge_tareo(ws, "B37:P37", "OBSERVACIONES:", bold=True, size=10, center=False, border=None)
    _merge_tareo(ws, "B38:P41", production.observations or "", size=9, center=False)
    _merge_tareo(ws, "C43:F43", "V°B° Jefe de Producción", bold=True, size=9)
    _merge_tareo(ws, "J43:M43", "V°B° Gerencia General", bold=True, size=9)
    _merge_tareo(ws, "B45:P45", "RESPONSABLE DEL REGISTRO", fill=TAREO_FOOTER, bold=True, size=9)
    _merge_tareo(ws, "D47:J47", "Generado por Control de Producción", size=9)
    _merge_tareo(ws, "M47:P47", crew_name, bold=True, size=9)
    for row in range(1, 48):
        ws.row_dimensions[row].height = 20
    ws.row_dimensions[2].height = 25
    ws.row_dimensions[3].height = 25
    ws.row_dimensions[4].height = 25
    ws.row_dimensions[5].height = 25
    ws.row_dimensions[13].height = 34
    ws.row_dimensions[21].height = 34
    ws.row_dimensions[29].height = 34
    ws.print_area = "B2:P47"


def _write_tareo_section(ws, row, title, crew_name, total_weight, count):
    _merge_tareo(ws, f"B{row}:P{row}", title, fill=TAREO_BLUE, bold=True, size=15, border="medium")
    header_row = row + 2
    detail_row = row + 3
    total_row = row + 5
    fields = [
        (f"B{header_row}", "N°", f"B{detail_row}", 1),
        (f"C{header_row}:J{header_row}", "APELLIDOS Y NOMBRES", f"C{detail_row}:J{detail_row}", crew_name),
        (f"K{header_row}:L{header_row}", "MENÚ / CENA", f"K{detail_row}:L{detail_row}", ""),
        (f"M{header_row}:N{header_row}", "TOTAL (PESO)", f"M{detail_row}:N{detail_row}", float(total_weight)),
        (f"O{header_row}:P{header_row}", "Dinos", f"O{detail_row}:P{detail_row}", int(count or 0)),
    ]
    for header_range, header, value_range, value in fields:
        if ":" in header_range:
            _merge_tareo(ws, header_range, header, fill=TAREO_HEADER, bold=True, size=10)
        else:
            _tareo_cell(ws[header_range], header, fill=TAREO_HEADER, bold=True, size=10)
        if ":" in value_range:
            _merge_tareo(ws, value_range, value, size=10, center=value_range.startswith(("B", "M", "O")))
        else:
            _tareo_cell(ws[value_range], value, size=10)
    ws[f"M{detail_row}"].number_format = '#,##0.00'
    _merge_tareo(ws, f"K{total_row}:L{total_row}", "TOTAL", bold=True, size=10)
    _merge_tareo(ws, f"M{total_row}:P{total_row}", float(total_weight), fill=TAREO_GREY, bold=True, size=12, border="medium")
    ws[f"M{total_row}"].number_format = '#,##0.00 "kg"'


def build_reception_report_pdf(production):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    from xml.sax.saxutils import escape

    detail_rows, crew_summary, cone_summary, total_weight = _build_rows(production)

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        rightMargin=16 * mm,
        leftMargin=16 * mm,
        topMargin=18 * mm,
        bottomMargin=16 * mm,
    )
    styles = getSampleStyleSheet()
    title = ParagraphStyle(
        "Title",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=18,
        textColor=colors.HexColor(f"#{FOREST}"),
    )
    subtitle = ParagraphStyle(
        "Subtitle",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=9,
        textColor=colors.HexColor(f"#{INK}"),
    )

    story = [
        Paragraph(escape(f"Tareo de recepción · PP {production.number}"), title),
        Paragraph(
            escape(f"{production.customer.name} · {production.plant_lot} · {production.reception_date.strftime('%d/%m/%Y')}"),
            subtitle,
        ),
        Spacer(1, 4 * mm),
    ]

    crew_table = [["Cuadrilla", "Registros", "Peso kg", "Cono pota kg"]]
    if crew_summary:
        for item in crew_summary:
            crew_table.append(
                [
                    item["crew"],
                    str(item["count"]),
                    f'{item["weight"]:,.2f}',
                    f'{item["cone_weight"]:,.2f}',
                ]
            )
    else:
        crew_table.append(["Sin registros", "-", "-", "-"])
    crew = Table(crew_table, colWidths=[66 * mm, 22 * mm, 28 * mm, 28 * mm], hAlign="LEFT")
    crew.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{FOREST}")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#D6DEDB")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor(f"#{PAPER}")]),
                ("ALIGN", (1, 1), (-1, -1), "CENTER"),
            ]
        )
    )
    story += [
        Paragraph("Resumen de cuadrillas", styles["Heading3"]),
        crew,
        Spacer(1, 4 * mm),
    ]

    cone_table = [["Cuadrilla", "Registros", "Peso kg"]]
    if cone_summary:
        for item in cone_summary:
            cone_table.append([item["crew"], str(item["count"]), f'{item["weight"]:,.2f}'])
    else:
        cone_table.append(["Sin registros", "-", "-"])
    cone = Table(cone_table, colWidths=[78 * mm, 22 * mm, 24 * mm], hAlign="LEFT")
    cone.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{FOREST}")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#D6DEDB")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor(f"#{PAPER}")]),
                ("ALIGN", (1, 1), (-1, -1), "CENTER"),
            ]
        )
    )
    story += [
        Paragraph("Peso de limpieza de cono de pota por cuadrilla", styles["Heading3"]),
        cone,
        Spacer(1, 4 * mm),
    ]

    detail_table = [["Fecha", "Vehículo", "Carro", "Producto", "Cuadrilla", "Dino", "Peso kg"]]
    for item in detail_rows:
        detail_table.append(
            [
                item["date"].strftime("%d/%m/%Y"),
                item["vehicle"],
                item["car_number"],
                item["product"],
                item["crew"],
                item["container"],
                f'{item["weight_kg"]:,.2f}',
            ]
        )
    detail = Table(
        detail_table,
        colWidths=[17 * mm, 23 * mm, 14 * mm, 52 * mm, 26 * mm, 13 * mm, 20 * mm],
        repeatRows=1,
        hAlign="LEFT",
    )
    detail.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{FOREST}")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#D6DEDB")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor(f"#{PAPER}")]),
                ("ALIGN", (1, 1), (-1, -1), "CENTER"),
            ]
        )
    )
    story += [
        Paragraph("Detalle de recepción", styles["Heading3"]),
        detail,
    ]

    doc.build(story)
    return buf.getvalue()
