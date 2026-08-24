from __future__ import annotations

from collections import defaultdict
from decimal import Decimal
from io import BytesIO
from copy import copy
from pathlib import Path

from django.conf import settings
from openpyxl.styles import Alignment
from openpyxl import load_workbook

from productions.models import (
    PlateEntry,
    ReceptionEntry,
    TunnelEntry,
)
from .crew_control import (
    _normalized_crew_name,
    _product_is_cone_pota,
    _tray_kg,
)


TEMPLATE_PATH = (
    Path(settings.BASE_DIR)
    / "productions"
    / "report_templates"
    / "FILETEROS-POTA_TAREO.xlsx"
)


def _first_reception_plate(production):
    entries = (
        ReceptionEntry.objects.filter(production=production, is_active=True)
        .select_related("vehicle")
        .order_by("created_at", "pk")
    )
    first = entries.filter(car_number="1").first() or entries.first()
    return first.vehicle.plate.upper() if first else "SIN REGISTRO"
CAR_COLUMNS = {
    1: ("C", "D"),
    2: ("E", "F"),
    3: ("G", "H"),
    4: ("I", "J"),
    5: ("K", "L"),
    6: ("M", "N"),
    7: ("O", "P"),
    8: ("Q", "R"),
    9: ("S", "T"),
}
MAX_DINO = 57


class ReceptionTareoReportError(Exception):
    pass


def _decimal(value):
    return Decimal(str(value or 0)).quantize(Decimal("0.01"))


def _spanish_day(value):
    names = (
        "Lunes",
        "Martes",
        "Miércoles",
        "Jueves",
        "Viernes",
        "Sábado",
        "Domingo",
    )
    return names[value.weekday()]


def _spanish_month(value):
    names = (
        "ENERO",
        "FEBRERO",
        "MARZO",
        "ABRIL",
        "MAYO",
        "JUNIO",
        "JULIO",
        "AGOSTO",
        "SEPTIEMBRE",
        "OCTUBRE",
        "NOVIEMBRE",
        "DICIEMBRE",
    )
    return f"{names[value.month - 1]} {value.year}"


def _responsible_name(entries):
    for entry in entries:
        name = entry.responsible.get_full_name().strip()
        if name:
            return name.upper()
        if entry.responsible.username:
            return entry.responsible.username.upper()
    return ""


def _cone_packaging_weight(production):
    """Peso total de las bandejas de conos de pota registradas en la producción.

    Solo suman las bandejas (tray_count) de cono de pota:
    - Bandejas colocadas en los racks de los túneles 1-6: TunnelEntry.
    - Bandejas en los plaqueros 1-3 (cada posición es una bachada): PlateEntry.

    Cada bandeja se convierte a kilogramos con la regla tray_kg. El tareo
    reparte el total en partes iguales entre las cuadrillas que muestra.
    """
    total = Decimal("0.00")
    tray_ratio = _tray_kg(production)

    for entry in list(
        TunnelEntry.objects.filter(production=production, is_active=True)
    ) + list(PlateEntry.objects.filter(production=production, is_active=True)):
        if not _product_is_cone_pota(entry.product):
            continue
        total += _decimal(entry.tray_count) * tray_ratio

    return total


def _clear_original_values(ws):
    # Keep the supplied format, formulas, merges, images and print settings. Only
    # remove the sample reception data that was included in the user's workbook.
    for row in range(18, 75):
        for column in range(3, 23):
            ws.cell(row=row, column=column).value = None
    for column in range(3, 23, 2):
        ws.cell(row=14, column=column).value = None
        ws.cell(row=15, column=column).value = None
        ws.cell(row=16, column=column).value = None
    for column in range(3, 23):
        ws.cell(row=17, column=column).value = None


def _style_reception_header(ws):
    """Normalize the employer header so LibreOffice/PDF keeps it centered."""
    cell = ws["B7"]
    font = copy(cell.font)
    font.name = "Arial"
    font.sz = 13
    font.bold = True
    font.color = "FF000000"
    cell.font = font
    cell.alignment = Alignment(
        horizontal="centerContinuous",
        vertical="center",
        wrap_text=False,
    )


def _style_reception_entry_cell(cell):
    """Use one reliable style for every value written in the reception grid.

    The supplied file mixes Arial Narrow 16, Calibri 11 and cells with no
    alignment.  LibreOffice preserves those differences in the PDF, so two
    weights from the same car can look like they have different sizes or are
    shifted to a side.  Only values entered by the application use this style;
    the official headings and the rest of the template stay untouched.
    """
    font = copy(cell.font)
    font.name = "Arial"
    font.sz = 13
    font.bold = False
    font.color = "FF000000"
    cell.font = font
    cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=False,
    )


def _style_crew_sheet_top(ws):
    """Normalize the visible crew block so PDF rendering keeps one size."""
    target_cells = (
        "C16",
        "C17",
        "C18",
        "K16",
        "K17",
        "K18",
        "M16",
        "M17",
        "M18",
        "O16",
        "O17",
        "O18",
    )
    for coordinate in target_cells:
        cell = ws[coordinate]
        font = copy(cell.font)
        font.name = "Arial"
        font.sz = 13
        cell.font = font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )


# Celda (rango combinado, celda destino) donde vive cada hora de la página.
# Los rangos cubren TODO el hueco visible entre las etiquetas de la fila 9:
#   - "HORA INICIO:" (B9:C9) -> el hueco hasta "HORA TERMINO:" es D9:I9
#   - "HORA TERMINO:" (J9:K9) -> el resto de la línea es L9:P9
HOUR_RANGES = {
    "CUADRILLA 1": {
        "start": ("D9:I9", "D9"),
        "end": ("L9:P9", "L9"),
    },
    "CUADRILLA 2": {
        "start": ("D9:I9", "D9"),
        "end": ("L9:P9", "L9"),
    },
}


def _combine_hours(ws, sheet_name, start_time, end_time):
    """Put each hour in a combined cell so the PDF renders it centered."""
    layout = HOUR_RANGES[sheet_name]
    for kind, value in (("start", start_time), ("end", end_time)):
        cell_range, anchor = layout[kind]
        if cell_range in ws.merged_cells.ranges:
            ws.unmerge_cells(cell_range)
        for coordinate in ws[cell_range]:
            for cell in coordinate:
                cell.value = None
        ws.merge_cells(cell_range)
        ws[anchor] = value
        if value:
            ws[anchor].number_format = "h:mm AM/PM"
            ws[anchor].alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=False,
            )


def _normalize_crew_typography(ws):
    """Apply one coherent typography to every filled page.

    The official workbook ships titles at 18pt, column headers at 11pt, field
    labels at 9-10pt and values at a mix of sizes, which is why the PDF and the
    Excel look inconsistent.  Classify each cell and give it a fixed size.
    Values that summarise a block (totals), the hours and the crew names stay
    bold and black; the rest of the numbers keep a lighter regular weight.
    """
    title_cells = {"E2", "B13", "B21", "B29"}
    header_cells = {
        "B15", "C15", "K15", "M15", "O15",
        "B23", "C23", "K23", "M23", "O23",
        "B31", "C31", "K31", "M31", "O31",
    }
    label_cells = {
        "M2", "M3", "M4", "M5",
        "B7", "J7", "B8", "F8", "J8", "B9", "J9", "B10", "J10",
        "K18", "K26", "K34", "K36",
        "B37", "C43", "J43",
        "B45", "B46", "K46", "B47", "K47", "D47",
    }
    strong_value_cells = {
        "D9", "L9",  # hours
        "L10",  # plate
        "C16", "C24", "C32",  # crew names
        "M18", "M26", "M34", "N36",  # totals
    }
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            coordinate = cell.coordinate
            font = copy(cell.font)
            font.name = "Arial"
            font.color = "FF000000"
            if coordinate == "E2":
                font.sz = 18
                font.bold = True
            elif coordinate in title_cells:
                font.sz = 16
                font.bold = True
            elif coordinate in header_cells:
                font.sz = 11
                font.bold = True
            elif coordinate in label_cells:
                font.sz = 10
                font.bold = True
            elif coordinate in strong_value_cells:
                font.sz = 13
                font.bold = True
            else:
                font.sz = 13
                font.bold = False
            cell.font = font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=False,
            )


def _fill_reception_sheet(ws, production, entries):
    _clear_original_values(ws)
    _style_reception_header(ws)
    ws["U4"] = f"{_spanish_month(production.reception_date).replace(' ', '-') }"
    ws["E10"] = production.customer.name
    ws["T10"] = production.customer_lot
    ws["E11"] = production.process
    ws["U11"] = production.reception_date
    ws["U11"].number_format = "dd/mm/yyyy"

    entries_by_car = defaultdict(list)
    for entry in entries:
        car_text = (entry.car_number or "").strip()
        dino_text = (entry.container or "").strip()
        if not car_text.isdigit() or int(car_text) not in CAR_COLUMNS:
            raise ReceptionTareoReportError(
                f"El carro {car_text or 'sin número'} no cabe en la plantilla de tareo (carros 1 al 9)."
            )
        if not dino_text.isdigit() or not 1 <= int(dino_text) <= MAX_DINO:
            raise ReceptionTareoReportError(
                f"El dino {dino_text or 'sin número'} no cabe en la plantilla de tareo (dinos 1 al {MAX_DINO})."
            )
        entries_by_car[int(car_text)].append(entry)

    for car_number, car_entries in sorted(entries_by_car.items()):
        left_column, right_column = CAR_COLUMNS[car_number]
        first = car_entries[0]
        for coordinate, value in (
            (f"{left_column}14", first.product.description.upper()),
            (f"{left_column}15", first.vehicle.plate.upper()),
            (f"{left_column}16", car_number),
        ):
            ws[coordinate] = value
            _style_reception_entry_cell(ws[coordinate])

        crew_names = []
        for entry in car_entries:
            crew_name = _normalized_crew_name(entry.crew.name) if entry.crew_id else "SIN CUADRILLA"
            if crew_name not in crew_names:
                crew_names.append(crew_name)
        if len(crew_names) > 2:
            raise ReceptionTareoReportError(
                f"El carro {car_number} tiene más de dos cuadrillas y la plantilla solo admite dos."
            )
        for column, crew_name in zip((left_column, right_column), crew_names):
            ws[f"{column}17"] = crew_name
            _style_reception_entry_cell(ws[f"{column}17"])

        crew_column = {
            crew_name: column
            for column, crew_name in zip((left_column, right_column), crew_names)
        }
        for entry in car_entries:
            crew_name = _normalized_crew_name(entry.crew.name) if entry.crew_id else "SIN CUADRILLA"
            row = 17 + int(entry.container)
            ws[f"{crew_column[crew_name]}{row}"] = float(_decimal(entry.weight_kg))
            _style_reception_entry_cell(ws[f"{crew_column[crew_name]}{row}"])


def _fill_crew_sheets(wb, production, entries):
    crew_first_seen = []
    for entry in entries:
        crew_name = _normalized_crew_name(entry.crew.name) if entry.crew_id else "SIN CUADRILLA"
        if crew_name not in crew_first_seen:
            crew_first_seen.append(crew_name)

    cone_total = _cone_packaging_weight(production)
    # The official workbook contains exactly two tareo pages. Use the first two
    # reception crews and preserve those two sheets unchanged.
    selected_crews = crew_first_seen[:2]
    while len(selected_crews) < 2:
        selected_crews.append("")
    active_sheet_count = sum(1 for name in selected_crews if name)
    cone_share = (
        round((cone_total / Decimal(active_sheet_count)), 2)
        if active_sheet_count and cone_total
        else Decimal("0.00")
    )

    time_values = [entry.time for entry in entries if entry.time]
    start_time = min(time_values) if time_values else None
    end_time = max(time_values) if time_values else None
    supervisor = _responsible_name(entries)
    first_plate = _first_reception_plate(production)

    for index, sheet_name in enumerate(("CUADRILLA 1", "CUADRILLA 2")):
        ws = wb[sheet_name]
        _style_crew_sheet_top(ws)
        crew_name = selected_crews[index]

        ws["O4"] = _spanish_month(production.reception_date)
        ws["D7"] = production.customer.name
        ws["M7"] = production.main_product.description.upper()
        ws["D8"] = _spanish_day(production.reception_date)
        ws["G8"] = production.reception_date
        ws["G8"].number_format = "dd/mm/yyyy"
        ws["M8"] = production.get_shift_display().upper()
        _combine_hours(ws, sheet_name, start_time, end_time)
        ws["D10"] = supervisor
        plate_range = "L10:P10"
        if plate_range in ws.merged_cells.ranges:
            ws.unmerge_cells(plate_range)
        ws.merge_cells(plate_range)
        ws["L10"] = first_plate
        ws["L10"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        # The cone of pota cleaning weight: the total envasado across all
        # tunnels and plates is shared equally between the tareo crews.
        ws["M32"] = float(cone_share)
        ws["D46"] = supervisor
        ws["M46"] = production.reception_date
        ws["M46"].number_format = "dd/mm/yyyy"
        # Normalize the whole page: titles, column headers, field labels and
        # every value (filled directly or through a formula) keep a single
        # readable size each, so the PDF and the Excel look alike.
        _normalize_crew_typography(ws)


def _repair_broken_template_metadata(wb):
    """Remove only stale names/links left by a missing COLORES worksheet."""
    for name in ("CODIGOS", "LISTA"):
        wb.defined_names.pop(name, None)
        for worksheet in wb.worksheets:
            worksheet.defined_names.pop(name, None)

    reception_sheet = wb["POTA ENTERA"]
    for validation in reception_sheet.data_validations.dataValidation:
        if validation.formula1 and "LISTA" in validation.formula1.upper():
            validation.formula1 = "$AB$7:$AB$8"


def build_reception_tareo_xlsx(production):
    if not TEMPLATE_PATH.exists():
        raise ReceptionTareoReportError("No se encontró la plantilla oficial de tareo.")

    entries = list(
        ReceptionEntry.objects.filter(production=production, is_active=True)
        .select_related("vehicle", "product", "crew", "responsible")
        .order_by("car_number", "container", "pk")
    )
    if not entries:
        raise ReceptionTareoReportError("Todavía no hay pesos de recepción para generar el tareo.")

    wb = load_workbook(TEMPLATE_PATH, keep_links=False)
    _repair_broken_template_metadata(wb)
    _fill_reception_sheet(wb["POTA ENTERA"], production, entries)
    _fill_crew_sheets(wb, production, entries)
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"

    payload = BytesIO()
    wb.save(payload)
    return payload.getvalue()
