"""Tareo de cuadrilla — Excel con plantilla oficial (SPM-PRO-FOT-007-TPE) + PDF."""

import os
import shutil
import subprocess
import tempfile
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from .crew_control import crew_control_summary as crew_tareo_summary


class CrewTareoReportError(Exception):
    pass


def _first_reception_plate(production):
    from productions.models import ReceptionEntry

    entries = (
        ReceptionEntry.objects.filter(production=production, is_active=True)
        .select_related("vehicle")
        .order_by("created_at", "pk")
    )
    first = entries.filter(car_number="1").first() or entries.first()
    return first.vehicle.plate.upper() if first else "SIN REGISTRO"


TEMPLATE_PATH = Path(__file__).resolve().parent.parent / "templates_excel" / "tareo_envasado.xlsx"


def _shift_label(production):
    from productions.models import ProductionOrder

    return ProductionOrder.Shift(production.shift).label if production.shift else ""


def _fill_cell(ws, coord, value):
    ws[coord] = value


def build_crew_tareo_xlsx(
    production,
    crew_pk,
    hora_inicio: str = "",
    hora_termino: str = "",
    supervisor: str = "",
) -> bytes:
    tareo = crew_tareo_summary(production, crew_pk)
    if tareo is None:
        raise CrewTareoReportError("La cuadrilla solicitada no existe.")
    if not TEMPLATE_PATH.is_file():
        raise CrewTareoReportError("No se encontró la plantilla oficial de tareo de envasado.")

    wb = load_workbook(TEMPLATE_PATH)
    ws = wb.active

    # DATOS DEL PARTE / CUADRILLA
    _fill_cell(ws, "E11", production.customer.name if production.customer_id else "")
    _fill_cell(ws, "L11", tareo["crew_name"])
    _fill_cell(ws, "E12", production.production_date.strftime("%d/%m/%Y"))
    _fill_cell(ws, "L12", _shift_label(production))
    _fill_cell(ws, "E13", hora_inicio.strip())
    _fill_cell(ws, "L13", hora_termino.strip())
    _fill_cell(ws, "E14", supervisor.strip())
    _fill_cell(ws, "L14", production.plant_lot or "")

    # TABLA DE PERSONAL
    workers = tareo["workers"]
    name_row = 17
    last_used_row = name_row
    if workers:
        for index, worker in enumerate(workers[:15], start=1):
            row = name_row + index - 1
            _fill_cell(ws, f"B{row}", index)
            _fill_cell(ws, f"C{row}", worker.full_name)
            last_used_row = row
    else:
        _fill_cell(ws, "C17", "Sin trabajadores registrados")

    # TOTAL
    _fill_cell(ws, "M32", float(tareo["total_kg"]))

    # OBSERVACIONES
    _fill_cell(
        ws,
        "C33",
        (
            f"PP {production.number} · Lote {production.plant_lot or ''} · "
            f"{tareo['total_trays']} bandejas de {tareo['total_kg']} kg "
            f"({tareo['tunnel_trays']} túneles / {tareo['plate_trays']} plaqueros)."
        ),
    )

    # RESUMEN POR PRODUCTO (placa del primer carro + producto principal)
    first_plate = _first_reception_plate(production)
    main_product = production.main_product
    product_label = (
        f"{main_product.code} · {main_product.description}" if main_product else ""
    )
    _fill_cell(ws, "C38", f"{first_plate} · {product_label}" if product_label else first_plate)
    _fill_cell(ws, "G38", float(tareo["total_kg"]))

    payload = BytesIO()
    wb.save(payload)
    return payload.getvalue()


def build_crew_tareo_pdf(
    production,
    crew_pk,
    hora_inicio: str = "",
    hora_termino: str = "",
    supervisor: str = "",
) -> bytes:
    xlsx_payload = build_crew_tareo_xlsx(production, crew_pk, hora_inicio, hora_termino, supervisor)

    binary = shutil.which("soffice") or shutil.which("libreoffice")
    if not binary:
        raise CrewTareoReportError(
            "No se encontró LibreOffice para convertir el tareo a PDF."
        )

    stem = f"CUADRILLA_TAREO_PP_{production.number}"
    with tempfile.TemporaryDirectory(prefix="crew-tareo-pdf-") as temp_name:
        temp_dir = Path(temp_name)
        profile_dir = temp_dir / "libreoffice-profile"
        profile_dir.mkdir()
        input_path = temp_dir / f"{stem}.xlsx"
        output_path = temp_dir / f"{stem}.pdf"
        input_path.write_bytes(xlsx_payload)

        result = subprocess.run(
            [
                binary,
                "--headless",
                "--nologo",
                "--nodefault",
                "--nofirststartwizard",
                f"-env:UserInstallation={profile_dir.resolve().as_uri()}",
                "--convert-to",
                "pdf:calc_pdf_Export",
                "--outdir",
                str(temp_dir),
                str(input_path),
            ],
            cwd=temp_dir,
            env=os.environ.copy(),
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            timeout=120,
            check=False,
        )
        if result.returncode != 0 or not output_path.is_file():
            stderr = result.stderr.decode("utf-8", errors="replace").strip()
            detail = stderr or "LibreOffice no generó el PDF."
            raise CrewTareoReportError(
                f"No se pudo convertir el tareo a PDF: {detail[:300]}"
            )

        payload = output_path.read_bytes()
        if not payload.startswith(b"%PDF"):
            raise CrewTareoReportError("El tareo generado no es un PDF válido.")
        return payload
